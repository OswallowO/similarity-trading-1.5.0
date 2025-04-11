#這個版本有刪除很多未呼叫的函數，請注意
#該完成的項目：整合下單程式
import fugle_marketdata as fg
import pandas as pd
import yaml
import json
import os
import numpy as np
import openpyxl
import math
import colorama
import subprocess
import sys
import time as time_module
import warnings
from tabulate import tabulate
from openpyxl.styles import PatternFill
from colorama import init, Fore, Style
from datetime import datetime, time, timedelta, date
from fugle_marketdata import RestClient
from fugle_realtime import WebSocketClient
from concurrent.futures import ThreadPoolExecutor, as_completed
import websocket
import threading
import msvcrt
import traceback
import shioaji as sj
import shioaji_logic
import importlib
import touchprice as tp

colorama.init(autoreset=True)
warnings.filterwarnings("ignore", category=FutureWarning)

required_packages = [
    'fugle-marketdata',
    'pandas',
    'pyyaml',
    'colorama',
    'numpy',
    'python-dateutil',
    'tabulate',
    'openpyxl'
]

def install_package(package):
    try:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", package],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        print(f"{package} 安裝成功")
    except subprocess.CalledProcessError:
        print(f"{package} 安裝失敗")

def check_and_install_packages(packages):
    for package in packages:
        try:
            __import__(package)
            print(f"{package} 已安裝")
        except ImportError:
            install_package(package)

init(autoreset=True)

RED = Fore.RED
GREEN = Fore.GREEN
YELLOW = Fore.YELLOW
BLUE = Fore.BLUE
RESET = Style.RESET_ALL

pd.set_option('future.no_silent_downcasting', True)

def init_fugle_client():
    try:
        config = load_config("config.yaml")
        client = RestClient(api_key=config['api_key'])
        print("=" * 50)
        print("從 config.yaml 載入 API 金鑰")
        print("=" * 50)
        return client, config['api_key']
    except FileNotFoundError:
        print("錯誤：config.yaml 文件不存在。")
        sys.exit(1)
    except KeyError:
        print("錯誤：config.yaml 中缺少 'api_key'。")
        sys.exit(1)
    except Exception as e:
        print(f"初始化富果API客戶端時發生錯誤：{e}")
        sys.exit(1)

def load_config(config_file):
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    except FileNotFoundError:
        print(f"錯誤：無法找到 {config_file} 文件。")
        sys.exit(1)
    except yaml.YAMLError as e:
        print(f"錯誤：讀取 {config_file} 文件時發生 YAML 錯誤：{e}")
        sys.exit(1)

def calculate_5min_pct_increase_and_highest(intraday_df):
    """
    修改後的計算方式：
    1. 第一根K棒（例如 09:00）的 5min_pct_increase 固定為 0。
    2. 第二到第四根K棒（例如 09:01~09:03）：取從第一根到當前所有K棒的 close 值，
         如果最後一根的 close >= 第一根的 close（上升趨勢），公式為
              (最大close - 最小close) * 100 / 最小close
         否則（下降趨勢），公式為
              (最小close - 最大close) * 100 / 最大close
    3. 從第五根K棒（9:04以後）開始，取最近5根K棒的 close 值，按上述相同方式計算。
    
    同時，每根K棒的 highest 設為從開盤到當前的最高 high 值。
    
    傳入的 intraday_df 必須包含 'time', 'close', 'high' 欄位，且已按時間排序。
    """
    # 保證依時間排序
    intraday_df = intraday_df.sort_values(by='time').reset_index(drop=True)

    pct_increases = []
    highest_vals = []
    current_high = 0.0

    for idx, row in intraday_df.iterrows():
        try:
            close_val = float(row['close'])
        except Exception:
            close_val = 0.0
        try:
            high_val = float(row.get('high', 0.0))
        except Exception:
            high_val = close_val
        
        # 累計當前最高 high 值
        current_high = max(current_high, high_val)
        highest_vals.append(current_high)

        if idx == 0:
            # 第一根K棒：預設為 0
            pct_increases.append(0.0)
        else:
            # 決定取幾根K棒：若不足5根則取 idx+1 根；若足夠則取最近5根（idx-4 至 idx）
            if idx < 4:
                start_idx = 0
            else:
                start_idx = idx - 4
            window = intraday_df.loc[start_idx: idx, 'close']
            try:
                close_values = window.astype(float).tolist()
            except Exception:
                close_values = []
            if len(close_values) < 2:
                pct_increases.append(0.0)
            else:
                first_close = close_values[0]
                last_close = close_values[-1]
                max_close = max(close_values)
                min_close = min(close_values)
                # 根據趨勢計算：若最後值大於等於第一值，視為上升趨勢；否則為下降趨勢
                if last_close >= first_close:
                    # 上升趨勢：公式 (最大 - 最小)*100 / 最小
                    pct = (max_close - min_close) * 100 / min_close if min_close != 0 else 0.0
                else:
                    # 下降趨勢：公式 (最小 - 最大)*100 / 最大，結果為負值
                    pct = (min_close - max_close) * 100 / max_close if max_close != 0 else 0.0
                pct_increases.append(pct)
    
    intraday_df['5min_pct_increase'] = pct_increases
    intraday_df['highest'] = highest_vals
    return intraday_df

def fetch_intraday_data(client, symbol, trading_day, yesterday_close_price, start_time=None, end_time=None):
    try:
        symbol = str(symbol).strip()
        if not symbol:
            print(f"無效的 symbol: {symbol}")
            return pd.DataFrame()

        if isinstance(trading_day, str):
            trading_day_date = datetime.strptime(trading_day, '%Y-%m-%d').date()
        elif isinstance(trading_day, datetime):
            trading_day_date = trading_day.date()
        elif isinstance(trading_day, date):
            trading_day_date = trading_day
        else:
            print(f"未知的 trading_day 類型：{type(trading_day)}，值：{trading_day}")
            return pd.DataFrame()

        today = datetime.now().date()
        if trading_day_date < today:
            end_time_str = "13:30"
        else:
            current_time = datetime.now()
            market_end_time = current_time.replace(hour=13, minute=30, second=0, microsecond=0)
            if current_time > market_end_time:
                end_time_str = "13:30"
            else:
                if current_time.second == 0 and current_time.microsecond == 0:
                    effective_time = current_time - timedelta(minutes=1)
                else:
                    effective_time = current_time.replace(second=0, microsecond=0)
                end_time_str = effective_time.strftime('%H:%M')

        if start_time:
            _from = f"{trading_day}T{start_time}:00+08:00"
        else:
            _from = f"{trading_day}T09:00:00+08:00"

        if end_time:
            to = f"{trading_day}T{end_time}:00+08:00"
        else:
            to = f"{trading_day}T{end_time_str}:00+08:00"

        try:
            pd.to_datetime(_from)
            pd.to_datetime(to)
        except Exception as e:
            print(f"日期時間格式錯誤：_from={_from}, to={to}, 錯誤訊息：{e}")
            return pd.DataFrame()

        candles_response = client.stock.intraday.candles(
            symbol=symbol,
            timeframe='1',
            _from=_from,
            to=to
        )
        if not candles_response or 'data' not in candles_response or not candles_response['data']:
            print(f"API 回應無數據或格式不正確：{candles_response}")
            return pd.DataFrame()

        candles = candles_response['data']
        candles_df = pd.DataFrame(candles)
        if 'date' in candles_df.columns:
            candles_df['datetime'] = pd.to_datetime(candles_df['date'], errors='coerce')
        else:
            print(f"API 回應缺少 'date' 欄位，無法進行日期時間轉換。")
            return pd.DataFrame()

        required_fields = ['datetime', 'open', 'high', 'low', 'close', 'volume']
        existing_fields = candles_df.columns.tolist()
        missing_fields = [field for field in required_fields if field not in existing_fields]
        if missing_fields:
            print(f"API 回應缺少必要欄位：{missing_fields}")

        candles_df['date'] = candles_df['datetime'].dt.strftime('%Y-%m-%d')
        candles_df['time'] = candles_df['datetime'].dt.strftime('%H:%M:%S')

        if not candles_df.empty:
            candles_df.set_index('datetime', inplace=True)
            full_datetime_index = pd.date_range(start=_from, end=to, freq='1min')
            candles_df = candles_df.reindex(full_datetime_index)
            candles_df.reset_index(inplace=True)
            candles_df.rename(columns={'index': 'datetime'}, inplace=True)
            candles_df['date'] = candles_df['datetime'].dt.strftime('%Y-%m-%d')
            candles_df['time'] = candles_df['datetime'].dt.strftime('%H:%M:%S')
            candles_df['symbol'] = symbol
            candles_df['昨日收盤價'] = yesterday_close_price
            candles_df['漲停價'] = calculate_limit_up_price(yesterday_close_price)
            candles_df[['symbol', '昨日收盤價', '漲停價']] = candles_df[['symbol', '昨日收盤價', '漲停價']].ffill().bfill()

            # 補齊價格欄位：若數據缺失則以昨日收盤價填充
            candles_df['close'] = candles_df['close'].ffill().fillna(candles_df['昨日收盤價'])
            candles_df['open'] = candles_df['open'].ffill().fillna(candles_df['close'])
            candles_df['low'] = candles_df['low'].ffill().fillna(candles_df['close'])
            # 修改 high 欄位的填充方式：僅用向前填充，若仍缺失則以 open 補上
            candles_df['high'] = candles_df['high'].ffill()
            candles_df['high'] = candles_df['high'].fillna(candles_df['open'])
            candles_df['volume'] = candles_df['volume'].fillna(0)
            # 計算 rise 欄位
            candles_df['rise'] = (candles_df['close'] - candles_df['昨日收盤價']) / candles_df['昨日收盤價'] * 100

            # 計算當日最高價：累計的 high 欄位最大值
            candles_df['highest'] = candles_df['high'].cummax()

            # 保留指定欄位順序
            candles_df = candles_df[[ 
                'symbol', 'date', 'time', 'open', 'high', 'low', 'close', 'volume',
                '昨日收盤價', '漲停價', 'rise', 'highest'
            ]]

            # 嘗試讀取現有的 auto_intraday.json（若存在且有該股票數據，取最後一筆的 close）
            previous_close = None
            if os.path.exists('auto_intraday.json'):
                try:
                    with open('auto_intraday.json', 'r', encoding='utf-8') as f:
                        existing_data = json.load(f)
                        if symbol in existing_data and len(existing_data[symbol]) > 0:
                            previous_close = existing_data[symbol][0]['close']
                except Exception as e:
                    print(f"讀取 auto_intraday.json 發生錯誤：{e}")

            # 補齊盤中 volume 為 0 的 K 棒
            for i in range(len(candles_df)):
                if pd.isna(candles_df.loc[i, 'volume']) or candles_df.loc[i, 'volume'] == 0:
                    if i > 0 and pd.notna(candles_df.loc[i-1, 'close']):
                        fill_val = candles_df.loc[i-1, 'close']
                    elif previous_close is not None:
                        fill_val = previous_close
                    else:
                        fill_val = yesterday_close_price
                    candles_df.loc[i, 'open'] = fill_val
                    candles_df.loc[i, 'high'] = fill_val
                    candles_df.loc[i, 'low'] = fill_val
                    candles_df.loc[i, 'close'] = fill_val

            # 確保 volume 欄位無 NaN
            candles_df['volume'] = candles_df['volume'].fillna(0)
            # 刪除多餘的 "average" 欄位
            if "average" in candles_df.columns:
                candles_df.drop(columns=["average"], inplace=True)

            return candles_df
        else:
            print(f"{symbol} 的一分K數據為空。")
            return pd.DataFrame()

    except Exception as e:
        print(f"取得即時K數據時發生錯誤：{e}")
        traceback.print_exc()
        return pd.DataFrame()

def get_recent_trading_day():
    today = datetime.now().date()
    current_time = datetime.now().time()
    market_close_time = datetime.strptime("13:30", "%H:%M").time()
    market_open_time = datetime.strptime("09:00", "%H:%M").time()
    
    def last_friday(date):
        while date.weekday() != 4:
            date -= timedelta(days=1)
        return date

    weekday = today.weekday()
    
    if weekday == 0:
        if current_time >= market_close_time:
            trading_day = today
        else:
            trading_day = last_friday(today)
    elif weekday == 5:
        trading_day = last_friday(today)
    elif weekday == 6:
        trading_day = last_friday(today)
    else:
        if current_time >= market_close_time:
            trading_day = today
        elif current_time < market_open_time:
            trading_day = today - timedelta(days=1)
            if trading_day.weekday() == 0:
                trading_day = last_friday(trading_day)
        else:
            trading_day = today
    return trading_day

def fetch_daily_kline_data(client, symbol, days=2):
    end_date = get_recent_trading_day()
    start_date = end_date - timedelta(days=days)
    start_date_str = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')

    print(f"正在取得 {symbol} 從 {start_date_str} 到 {end_date_str} 的日K數據...")

    try:
        data = client.stock.historical.candles(symbol=symbol, from_=start_date_str, to=end_date_str)
        if 'data' in data and data['data']:
            daily_kline_df = pd.DataFrame(data['data'])
            return daily_kline_df
        else:
            print(f"無法取得 {symbol} 的日K數據：API 回應中不包含 'data' 欄位或 'data' 為空")
            return pd.DataFrame()
    except Exception as e:
        print(f"無法取得 {symbol} 的日K數據：{e}")
        return pd.DataFrame()

def save_matrix_dict(matrix_dict):
    with open('matrix_dict_analysis.json', 'w', encoding='utf-8') as f:
        json.dump(matrix_dict, f, indent=4, ensure_ascii=False)

def load_matrix_dict_analysis():
    if os.path.exists('matrix_dict_analysis.json'):
        with open('matrix_dict_analysis.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        print("matrix_dict_analysis.json 文件不存在。")
        return {}

def save_nb_matrix_dict(nb_matrix_dict):
    with open('nb_matrix_dict.json', 'w', encoding='utf-8') as f:
        json.dump(nb_matrix_dict, f, indent=4, ensure_ascii=False, default=str)

def initialize_stock_data(symbols_to_analyze, daily_kline_data, intraday_kline_data):
    stock_data_collection = {}
    for symbol in symbols_to_analyze:
        if symbol not in daily_kline_data or symbol not in intraday_kline_data:
            print(f"股票代號 {symbol} 的日 K 線或一分 K 線資料缺失，跳過。")
            continue
        daily_kline_df = pd.DataFrame(daily_kline_data[symbol])
        intraday_data = pd.DataFrame(intraday_kline_data[symbol])
        if intraday_data.empty:
            print(f"股票代號 {symbol} 的日內數據為空，跳過。")
            continue
        complete_df = ensure_continuous_time_series(intraday_data)
        complete_df = complete_df.drop(columns=['volume', 'average'], errors='ignore')
        stock_data_collection[symbol] = complete_df
    return stock_data_collection

def process_group_data(stock_data_collection, wait_minutes, hold_minutes, matrix_dict_analysis, verbose=True):
    global capital_per_stock, transaction_fee, transaction_discount, trading_tax
    global price_gap_below_50, price_gap_50_to_100, price_gap_100_to_500, price_gap_500_to_1000, price_gap_above_1000
    global allow_reentry_after_stop_loss
    global in_position, has_exited, current_position, stop_loss_triggered
    global final_check_active, final_check_count, in_waiting_period, waiting_time
    global hold_time, leader, previous_rise_values

    merged_df = None
    total_profit = 0
    total_profit_rate = 0
    total_trades = 0
    message_log = []
    already_entered_stocks = []
    final_check_active = False
    final_check_count = 0
    final_check_max = 10
    can_trade = True
    already_triggered_limit_up = set()
    leader = None
    tracking_stocks = set()
    leader_rise_before_decline = None
    in_waiting_period = False
    waiting_time = 0
    hold_time = 0
    first_condition_one_time = None
    leader_peak_rise = None
    backtrack = False
    in_position = False
    has_exited = False
    current_position = None
    stop_loss_triggered = False
    previous_rise_values = {}
    group_name = None

    pull_up_entry = False
    limit_up_entry = False

    def truncate_to_two_decimals(value):
        return math.floor(value * 100) / 100 if value is not None else None

    for symbol, df in stock_data_collection.items():
        if not isinstance(df, pd.DataFrame):
            print(f"股票代號 {symbol} 的數據不是 DataFrame，跳過。")
            continue
        required_columns = ['time', 'rise', 'high', '漲停價', 'close', '5min_pct_increase']
        if not all(col in df.columns for col in required_columns):
            missing_cols = [col for col in required_columns if col not in df.columns]
            print(f"股票代號 {symbol} 的資料缺少必要列 {missing_cols}，跳過。")
            continue
        df_selected = df[['time', 'rise', 'high', '漲停價', 'close', '5min_pct_increase']].copy()
        df_selected = df_selected.rename(columns={
            'rise': f'rise_{symbol}',
            'high': f'high_{symbol}',
            '漲停價': f'limit_up_price_{symbol}',
            'close': f'close_{symbol}',
            '5min_pct_increase': f'5min_pct_increase_{symbol}'
        })

        if merged_df is None:
            merged_df = df_selected
        else:
            merged_df = pd.merge(merged_df, df_selected, on='time', how='outer')

    if merged_df is not None and not merged_df.empty:
        merged_df = merged_df.sort_values('time').reset_index(drop=True)
    else:
        merged_df = pd.DataFrame()

    total_bars = len(merged_df)
    merged_df_list = list(merged_df.iterrows())
    idx = 0
    stock_symbols = list(stock_data_collection.keys())

    def check_5min_pct_increase(stock, start_time, end_time):
        stock_df = stock_data_collection.get(stock, pd.DataFrame())
        if stock_df.empty:
            return False
        period_data = stock_df[(stock_df['time'] >= start_time) & (stock_df['time'] <= end_time)]
        return (period_data['5min_pct_increase'] >= 2).any()

    def check_high_values_during_period(stock, start_time, end_time):
        stock_df = stock_data_collection.get(stock, pd.DataFrame())
        if stock_df.empty:
            return False
        period_data = stock_df[(stock_df['time'] >= start_time) & (stock_df['time'] <= end_time)]
        period_data = period_data.sort_values(by='time').reset_index(drop=True)
        for i in range(1, len(period_data)):
            if period_data.loc[i, 'high'] <= period_data.loc[i - 1, 'high']:
                return True
        return False

    while idx < total_bars:
        index, row = merged_df_list[idx]
        current_time = row['time']
        current_time_str = current_time.strftime('%H:%M:%S')

        if current_time_str == '13:30:00' and in_position:
            profit, profit_rate = exit_trade(
                stock_data_collection[current_position['symbol']],
                current_position['shares'],
                current_position['entry_price'],
                current_position['sell_cost'],
                current_position['entry_fee'],
                current_position['tax'],
                message_log,
                current_time,
                hold_time,
                current_position['entry_time'],
                use_f_exit=True
            )
            if profit is not None and profit_rate is not None:
                total_trades += 1
                total_profit += profit
                total_profit_rate += profit_rate
            in_position = False
            has_exited = True
            current_position = None
            idx += 1
            continue

        for symbol in stock_symbols:
            stock_df = stock_data_collection[symbol]
            current_row = stock_df[stock_df['time'] == current_time]
            if not current_row.empty:
                rise_col = f'rise_{symbol}'
                row[rise_col] = current_row['rise'].values[0]
                high_col = f'high_{symbol}'
                row[high_col] = current_row['high'].values[0]
                row[f'5min_pct_increase_{symbol}'] = current_row['5min_pct_increase'].values[0]
                limit_up_price_col = f'limit_up_price_{symbol}'
                row[f'limit_up_price_{symbol}'] = truncate_to_two_decimals(current_row['漲停價'].values[0])
                close_col = f'close_{symbol}'
                row[f'close_{symbol}'] = current_row['close'].values[0]
            else:
                row[f'rise_{symbol}'] = None
                row[f'high_{symbol}'] = None
                row[f'5min_pct_increase_{symbol}'] = None
                row[f'limit_up_price_{symbol}'] = None
                row[f'close_{symbol}'] = None

        if in_position and not has_exited:
            hold_time += 1
            if hold_minutes is not None:
                if hold_time >= hold_minutes:
                    profit, profit_rate = exit_trade(
                        stock_data_collection[current_position['symbol']],
                        current_position['shares'],
                        current_position['entry_price'],
                        current_position['sell_cost'],
                        current_position['entry_fee'],
                        current_position['tax'],
                        message_log,
                        current_time,
                        hold_time,
                        current_position['entry_time']
                    )
                    if profit is not None and profit_rate is not None:
                        total_trades += 1
                        total_profit += profit
                        total_profit_rate += profit_rate
                    in_position = False
                    has_exited = True

            selected_symbol = current_position['symbol']
            selected_stock_df = stock_data_collection[selected_symbol]
            current_row = selected_stock_df[selected_stock_df['time'] == current_time]
            if not current_row.empty:
                current_high = current_row['high'].values[0]
                current_high_truncated = truncate_to_two_decimals(current_high)
                price_difference = (current_position['highest_on_entry'] - current_position['entry_price']) * 1000
                if price_difference < current_position['current_price_gap']:
                    stop_loss_type = 'price_difference'
                    stop_loss_threshold = current_position['entry_price'] + (current_position['current_price_gap'] / 1000)
                else:
                    stop_loss_type = 'over_high'
                    stop_loss_threshold = current_position['highest_on_entry'] + current_position['tick_unit']

                if current_high_truncated >= stop_loss_threshold:
                    exit_price = stop_loss_threshold
                    exit_reason = f"條件三觸發（{stop_loss_type}停損）"
                    trigger_exit = True
                else:
                    trigger_exit = False

                if trigger_exit:
                    exit_cost = current_position['shares'] * exit_price * 1000
                    exit_fee = int(exit_cost * (transaction_fee * 0.01) * (transaction_discount * 0.01))
                    profit = current_position['sell_cost'] - exit_cost - current_position['entry_fee'] - exit_fee - current_position['tax']
                    return_rate = (profit * 100) / (current_position['sell_cost'] - current_position['entry_fee'] - exit_fee) if (current_position['sell_cost'] - current_position['entry_fee'] - exit_fee) != 0 else 0.0
                    message_log.append(
                        (current_time_str,
                         f"{RED}{exit_reason}！出場成功！{RESET}")
                    )
                    message_log.append(
                        (current_time_str,
                         f"{RED}股票代號：{current_position['symbol']}，持有張數：{current_position['shares']} 張，"
                         f"出場價格：{exit_price} 元，出場價金：{int(exit_cost)} 元，利潤：{int(profit)} 元，"
                         f"報酬率：{return_rate:.2f}%，手續費：{exit_fee} 元{RESET}")
                    )
                    total_trades += 1
                    total_profit += profit
                    total_profit_rate += return_rate
                    in_position = False
                    has_exited = True
                    current_position = None
                    stop_loss_triggered = True

                    if allow_reentry_after_stop_loss:
                        backtrack_start_idx = max(0, idx - 5)
                        idx = backtrack_start_idx
                        backtrack = True
                        leader = None
                        tracking_stocks = set()
                        previous_rise_values.clear()
                        leader_rise_before_decline = None
                        in_waiting_period = False
                        waiting_time = 0
                        already_entered_stocks = []
                        final_check_active = False
                        final_check_count = 0
                        can_trade = True
                        hold_time = 0
                        first_condition_one_time = None
                        leader_peak_rise = None
                        pull_up_entry = False
                        limit_up_entry = False
                        if verbose:
                            try:
                                previous_time_str = merged_df_list[idx][1]['time'].strftime('%H:%M:%S')
                            except IndexError:
                                previous_time_str = "未知時間"
                            message_log.append(
                                (current_time_str, f"{YELLOW}[回朔] 觸發條件三，回溯五根K棒至 {previous_time_str}，檢查是否有新的進場機會{RESET}")
                            )
                        continue
                    else:
                        message_log.append((current_time_str, "停損後無其它進場機會，結束程序"))
                        break
            if in_position and not has_exited:
                idx += 1
                continue

        for symbol in stock_symbols:
            stock_df = stock_data_collection[symbol]
            current_row = stock_df[stock_df['time'] == current_time]
            if current_row.empty:
                continue

            limit_up_price = row.get(f'limit_up_price_{symbol}', None)
            current_high = row.get(f'high_{symbol}', None)
            pct_increase = row.get(f'5min_pct_increase_{symbol}', None)
            rise = row.get(f'rise_{symbol}', None)

            if current_high is None or pct_increase is None or rise is None:
                continue

            current_high_truncated = truncate_to_two_decimals(current_high) if current_high is not None else None
            limit_up_price_truncated = truncate_to_two_decimals(limit_up_price) if limit_up_price is not None else None

            if (current_time_str == '09:00:00' and current_high_truncated == limit_up_price_truncated) or \
               (current_time_str != '09:00:00' and current_high_truncated == limit_up_price_truncated and idx > 0 and \
                truncate_to_two_decimals(merged_df_list[idx - 1][1].get(f'high_{symbol}', None)) < limit_up_price_truncated):

                if pull_up_entry:
                    in_waiting_period = False
                    waiting_time = 0
                    final_check_active = False
                    final_check_count = 0
                    first_condition_one_time = None
                    leader_rise_before_decline = None
                    leader_peak_rise = None
                    tracking_stocks.clear()
                    previous_rise_values.clear()
                    leader = None
                    pull_up_entry = False
                    if verbose:
                        message_log.append(
                            (current_time_str, "觸發漲停進場，終止拉高進場的檢查")
                        )

                leader, in_waiting_period, waiting_time = limit_up_entry_function(
                    symbol, current_time, current_time_str, tracking_stocks, leader, in_waiting_period, waiting_time, message_log, verbose
                )
                pull_up_entry = False
                limit_up_entry = True
                leader_rise_before_decline = None
                leader_peak_rise = None
                previous_rise_values.clear()
                break

            if current_high_truncated != limit_up_price_truncated and pct_increase >= 2 and symbol not in tracking_stocks:
                first_condition_one_time = pull_up_entry_function(
                    symbol, current_time, current_time_str, row, message_log, tracking_stocks, verbose, final_check_active, in_waiting_period
                )
                pull_up_entry = True
                limit_up_entry = False

        if tracking_stocks:
            max_rise = None
            new_leader = leader
            for symbol in tracking_stocks:
                rise = row.get(f'rise_{symbol}', None)
                if rise is not None:
                    if max_rise is None or rise > max_rise:
                        max_rise = rise
                        new_leader = symbol

            if new_leader != leader or (leader_peak_rise is not None and max_rise > leader_peak_rise):
                if verbose and leader is not None:
                    message_log.append(
                        (current_time_str, f"領漲者變更為 {new_leader}，rise: {max_rise:.2f}%")
                    )
                leader = new_leader
                leader_peak_rise = max_rise
                leader_rise_before_decline = None

                if in_waiting_period and pull_up_entry:
                    in_waiting_period = False
                    waiting_time = 0
                    if verbose:
                        message_log.append(
                            (current_time_str, f"領漲變更，重置等待時間")
                        )

            if leader and not in_waiting_period and not final_check_active:
                rise = row.get(f'rise_{leader}', None)
                if verbose and rise is not None:
                    message_log.append(
                        (current_time_str, f"領漲 {leader}，rise: {rise:.2f}%")
                    )

            current_rise = row.get(f'rise_{leader}', None)
            prev_rise = previous_rise_values.get(leader)

            if not final_check_active and pull_up_entry:
                if prev_rise is not None and current_rise is not None:
                    if current_rise <= prev_rise:
                        if not in_waiting_period:
                            in_waiting_period = True
                            waiting_time = 1
                            previous_time = (datetime.combine(datetime.today(), current_time) - timedelta(minutes=1)).time()
                            previous_rise_value_series = stock_data_collection[leader][stock_data_collection[leader]['time'] == previous_time]['rise']
                            if not previous_rise_value_series.empty:
                                leader_rise_before_decline = previous_rise_value_series.values[0]
                            else:
                                leader_rise_before_decline = current_rise
                if leader and pull_up_entry:
                    previous_rise_values[leader] = current_rise

        if in_waiting_period:
            if limit_up_entry:
                if verbose:
                    message_log.append(
                        (current_time_str,
                         f"等待中，第 {waiting_time} 分鐘")
                    )
                if waiting_time >= wait_minutes:
                    in_waiting_period = False
                    waiting_time = 0
                    final_check_active = False
                    final_check_count = 0
                    if verbose:
                        message_log.append(
                            (current_time_str,
                             "等待完成，開始檢查是否有符合進場條件的股票")
                        )
                    eligible_stocks = []
                    group_name = None
                    for group, symbols in matrix_dict_analysis.items():
                        if leader in symbols:
                            group_name = group
                            break

                    if group_name is None:
                        print(f"無法找到領漲 {leader} 所屬的族群，無法進行檢查。")
                        idx += 1
                        continue

                    nb_matrix_dict = load_nb_matrix_dict()
                    consolidated_symbols = nb_matrix_dict.get('consolidated_symbols', {})
                    if group_name in consolidated_symbols:
                        nb_symbols = consolidated_symbols[group_name]
                        for symbol in nb_symbols:
                            if symbol == leader:
                                continue
                            stock_df = stock_data_collection.get(symbol, pd.DataFrame())
                            if stock_df.empty:
                                continue

                            has_pct_increase = check_5min_pct_increase(symbol, first_condition_one_time, current_time)
                            if not has_pct_increase:
                                continue

                            has_high_decrease = check_high_values_during_period(symbol, first_condition_one_time, current_time)
                            if not has_high_decrease:
                                continue

                            current_rise = row.get(f'rise_{symbol}', 0)
                            if not (-3 < current_rise < 8):
                                continue

                            eligible_stocks.append({
                                'symbol': symbol,
                                'rise': current_rise
                            })
                    else:
                        print(f"{group_name} 不在 consolidated_symbols 中")

                    if eligible_stocks:
                        entry_trade(
                            eligible_stocks, current_time, current_time_str, stock_data_collection, idx,
                            message_log, already_entered_stocks, tracking_stocks, previous_rise_values, verbose=verbose
                        )
                        pull_up_entry = False
                        limit_up_entry = False
                        idx += 1
                        continue
                    else:
                        if not final_check_active:
                            final_check_active = True
                            final_check_count = 0
                            if verbose:
                                message_log.append(
                                    (current_time_str,
                                         "沒有符合進場條件的股票，進入最後十次檢查階段")
                                )
                else:
                    waiting_time += 1
                idx += 1
                continue
            elif pull_up_entry:
                other_symbols = tracking_stocks - {leader} if leader else tracking_stocks
                if not other_symbols:
                    if verbose:
                        message_log.append(
                            (current_time_str, "等待中，僅有領漲股票，跳過重置流程")
                        )
                else:
                    for symbol in other_symbols.copy():
                        rise = row.get(f'rise_{symbol}', None)
                        if rise is not None and leader_rise_before_decline is not None:
                            if rise > leader_rise_before_decline:
                                final_check_active = False
                                final_check_count = 0
                                in_waiting_period = False
                                waiting_time = 0
                                leader_peak_rise = rise
                                if verbose:
                                    message_log.append(
                                        (current_time_str, f"領漲 {leader} 超越記錄的 rise 值，重置流程")
                                    )
                        elif symbol != leader:
                            leader_rise_before_decline = rise

                if in_waiting_period:
                    if verbose:
                        message_log.append(
                            (current_time_str,
                                 f"等待中，第 {waiting_time} 分鐘")
                        )
                    if waiting_time >= wait_minutes:
                        in_waiting_period = False
                        waiting_time = 0
                        final_check_active = False
                        final_check_count = 0
                        if verbose:
                            message_log.append(
                                (current_time_str,
                                     "等待完成，開始檢查是否有符合進場條件的股票")
                            )
                        eligible_stocks = []
                        group_name = None
                        for group, symbols in matrix_dict_analysis.items():
                            if leader in symbols:
                                group_name = group
                                break

                        if group_name is None:
                            print(f"無法找到領漲 {leader} 所屬的族群，無法進行檢查。")
                            idx += 1
                            continue

                        nb_matrix_dict = load_nb_matrix_dict()
                        consolidated_symbols = nb_matrix_dict.get('consolidated_symbols', {})
                        if group_name in consolidated_symbols:
                            nb_symbols = consolidated_symbols[group_name]
                            for symbol in nb_symbols:
                                if symbol == leader:
                                    continue
                                stock_df = stock_data_collection.get(symbol, pd.DataFrame())
                                if stock_df.empty:
                                    continue

                                has_pct_increase = check_5min_pct_increase(symbol, first_condition_one_time, current_time)
                                if not has_pct_increase:
                                    continue

                                has_high_decrease = check_high_values_during_period(symbol, first_condition_one_time, current_time)
                                if not has_high_decrease:
                                    continue

                                current_rise = row.get(f'rise_{symbol}', 0)
                                if not (-3 < current_rise < 8):
                                    continue

                                eligible_stocks.append({
                                    'symbol': symbol,
                                    'rise': current_rise
                                })
                        else:
                            print(f"{group_name} 不在 consolidated_symbols 中")

                        if eligible_stocks:
                            entry_trade(
                                eligible_stocks, current_time, current_time_str, stock_data_collection, idx,
                                message_log, already_entered_stocks, tracking_stocks, previous_rise_values, verbose=verbose
                            )
                            pull_up_entry = False
                            limit_up_entry = False
                            idx += 1
                            continue
                        else:
                            if not final_check_active:
                                final_check_active = True
                                final_check_count = 0
                                if verbose:
                                    message_log.append(
                                        (current_time_str,
                                             "沒有符合進場條件的股票，進入最後十次檢查階段")
                                    )
                    else:
                        waiting_time += 1
                idx += 1
                continue

        if final_check_active:
            if final_check_count >= final_check_max:
                if verbose:
                    message_log.append(
                        (current_time_str,
                        f"{YELLOW}最後檢查完成，仍未發現可進場股票{RESET}")
                    )

                final_check_active = False
                final_check_count = 0
                in_waiting_period = False
                waiting_time = 0
                hold_time = 0
                leader = None
                tracking_stocks.clear()
                previous_rise_values.clear()
                leader_peak_rise = None
                leader_rise_before_decline = None
                first_condition_one_time = None
                pull_up_entry = False
                limit_up_entry = False
                idx += 1
                continue

            final_check_count += 1
            if verbose:
                message_log.append(
                    (current_time_str,
                    f"最後檢查第 {final_check_count} 分鐘")
                )

            if leader and row.get(f'high_{leader}', None) == row.get(f'limit_up_price_{leader}', None):
                idx += 1
                continue

            if pull_up_entry:
                rise = row.get(f'rise_{leader}', None)
                if rise is not None and leader_rise_before_decline is not None and rise > leader_rise_before_decline:
                    final_check_active = False
                    final_check_count = 0
                    in_waiting_period = False
                    waiting_time = 0
                    leader_peak_rise = rise
                    if verbose:
                        message_log.append(
                            (current_time_str, f"領漲 {leader} 超越記錄的 rise 值，重置流程")
                        )
                    idx += 1
                    continue

            eligible_stocks = []
            group_name = None
            for group, symbols in matrix_dict_analysis.items():
                if leader in symbols:
                    group_name = group
                    break

            if group_name is None:
                print(f"無法找到領漲 {leader} 所屬的族群，無法進行檢查。")
                idx += 1
                continue

            nb_matrix_dict = load_nb_matrix_dict()
            consolidated_symbols = nb_matrix_dict.get('consolidated_symbols', {})
            if group_name in consolidated_symbols:
                nb_symbols = consolidated_symbols[group_name]
                for symbol in nb_symbols:
                    if symbol == leader:
                        continue

                    stock_df = stock_data_collection.get(symbol, pd.DataFrame())
                    if stock_df.empty:
                        continue

                    has_pct_increase = check_5min_pct_increase(symbol, first_condition_one_time, current_time)
                    if not has_pct_increase:
                        continue

                    has_high_decrease = check_high_values_during_period(symbol, first_condition_one_time, current_time)
                    if not has_high_decrease:
                        continue

                    current_rise = row.get(f'rise_{symbol}', 0)
                    if not (-3 < current_rise < 8):
                        continue

                    eligible_stocks.append({
                        'symbol': symbol,
                        'rise': current_rise
                    })
            else:
                print(f"{group_name} 不在 consolidated_symbols 中")

            if eligible_stocks:
                entry_trade(
                    eligible_stocks, current_time, current_time_str, stock_data_collection, idx,
                    message_log, already_entered_stocks, tracking_stocks, previous_rise_values, verbose=verbose
                )
                pull_up_entry = False
                limit_up_entry = False
                idx += 1
                continue
            else:
                if final_check_count >= final_check_max:
                    if verbose:
                        message_log.append(
                            (current_time_str,
                                 f"{YELLOW}最後檢查完成，仍未發現可進場股票{RESET}")
                        )

                    final_check_active = False
                    final_check_count = 0
                    in_waiting_period = False
                    waiting_time = 0
                    hold_time = 0
                    leader = None
                    tracking_stocks.clear()
                    previous_rise_values.clear()
                    leader_peak_rise = None
                    leader_rise_before_decline = None
                    first_condition_one_time = None
                    pull_up_entry = False
                    limit_up_entry = False
                idx += 1
                continue

        idx += 1

    message_log.sort(key=lambda x: str(x[0]))
    for log_time, message in message_log:
        print(f"[{log_time}] {message}")

    if total_trades > 0:
        avg_profit_rate = total_profit_rate / total_trades
        print(f"\n族群 {group_name} 的模擬交易完成，總利潤：{int(total_profit)} 元，平均報酬率：{avg_profit_rate:.2f}%\n")
        return total_profit, avg_profit_rate
    else:
        if verbose:
            print("無交易，無法計算總利潤和報酬率")
        return None, None

def pull_up_entry_function(symbol, current_time, current_time_str, row, message_log, tracking_stocks, verbose=True, final_check_active=False, in_waiting_period=False):
    global pull_up_entry, limit_up_entry
    if symbol not in tracking_stocks:
        tracking_stocks.add(symbol)
        if verbose and not in_waiting_period and not final_check_active:
            message_log.append(
                (current_time_str, f"股票代號:{symbol} 觸發拉高進場條件")
            )
    first_condition_one_time = current_time
    pull_up_entry = True
    limit_up_entry = False
    return first_condition_one_time

def limit_up_entry_function(symbol, current_time, current_time_str, tracking_stocks, leader, in_waiting_period, waiting_time, message_log, verbose=True):
    global pull_up_entry, limit_up_entry
    tracking_stocks.clear()
    tracking_stocks.add(symbol)
    leader = symbol
    in_waiting_period = True
    waiting_time = 1
    pull_up_entry = False
    limit_up_entry = True
    if verbose:
        message_log.append(
            (current_time_str, f"領漲 {symbol} 漲停，觸發漲停進場條件")
        )
    return leader, in_waiting_period, waiting_time
    
def entry_trade(
    eligible_stocks, current_time, current_time_str, stock_data_collection, idx,
    message_log, already_entered_stocks, tracking_stocks, previous_rise_values, verbose=True
):
    global capital_per_stock, transaction_fee, transaction_discount, trading_tax
    global price_gap_below_50, price_gap_50_to_100, price_gap_100_to_500
    global price_gap_500_to_1000, price_gap_above_1000
    global in_position, has_exited, current_position
    global allow_reentry_after_stop_loss, stop_loss_triggered
    global final_check_active, final_check_count, in_waiting_period, waiting_time
    global hold_time, leader

    if in_position:
        if verbose:
            message_log.append(
                (current_time_str, f"{YELLOW}已有持倉，無法進行新的進場操作{RESET}")
            )
        return

    eligible_stocks_sorted = sorted(eligible_stocks, key=lambda x: x['rise'], reverse=True)
    median_index = len(eligible_stocks_sorted) // 2
    selected_stock = eligible_stocks_sorted[median_index]
    selected_symbol = selected_stock['symbol']
    selected_stock_rise = selected_stock['rise']
    
    entry_price_series = stock_data_collection[selected_symbol][stock_data_collection[selected_symbol]['time'] == current_time]['close']
    if not entry_price_series.empty:
        entry_price = entry_price_series.values[0]
        shares = round((capital_per_stock * 10000) / (entry_price * 1000))
        sell_cost = shares * entry_price * 1000
        entry_fee = int(sell_cost * (transaction_fee * 0.01) * (transaction_discount * 0.01))
        tax = int(sell_cost * (trading_tax * 0.01))
        
        if entry_price < 10:
            current_price_gap = price_gap_below_50
            tick_unit = 0.01
        elif entry_price < 50:
            current_price_gap = price_gap_50_to_100
            tick_unit = 0.05
        elif entry_price < 100:
            current_price_gap = price_gap_50_to_100
            tick_unit = 0.1
        elif entry_price < 500:
            current_price_gap = price_gap_100_to_500
            tick_unit = 0.5
        elif entry_price < 1000:
            current_price_gap = price_gap_500_to_1000
            tick_unit = 1
        else:
            current_price_gap = price_gap_above_1000
            tick_unit = 5

        highest_on_entry_series = stock_data_collection[selected_symbol][stock_data_collection[selected_symbol]['time'] == current_time]['high']
        if not highest_on_entry_series.empty:
            highest_on_entry = highest_on_entry_series.values[0]
        else:
            highest_on_entry = entry_price

        current_position = {
            'symbol': selected_symbol,
            'shares': shares,
            'entry_price': entry_price,
            'sell_cost': sell_cost,
            'entry_fee': entry_fee,
            'tax': tax,
            'entry_time': current_time_str,
            'entry_index': idx,
            'current_price_gap': current_price_gap,
            'tick_unit': tick_unit,
            'highest_on_entry': highest_on_entry,
            'initial_highest': highest_on_entry,
            'stop_loss_type': None,
            'stop_loss_threshold': None
        }
        message_log.append(
            (current_time_str,
             f"{GREEN}進場！股票代號：{selected_symbol}，進場 {shares} 張，進場價格：{entry_price} 元，"
             f"進場價金：{int(sell_cost)} 元，手續費：{entry_fee} 元，證交稅：{tax} 元。{RESET}")
        )

        in_position = True
        has_exited = False
        already_entered_stocks.append(selected_symbol)
        hold_time = 0

        if allow_reentry_after_stop_loss:
            stop_loss_triggered = False

        price_difference = (current_position['highest_on_entry'] - current_position['entry_price']) * 1000
        if price_difference < current_position['current_price_gap']:
            current_position['stop_loss_type'] = 'price_difference'
            current_position['stop_loss_threshold'] = current_position['entry_price'] + (current_position['current_price_gap'] / 1000)
        else:
            current_position['stop_loss_type'] = 'over_high'
            current_position['stop_loss_threshold'] = current_position['highest_on_entry'] + current_position['tick_unit']

        final_check_active = False
        final_check_count = 0
        in_waiting_period = False
        waiting_time = 0
        hold_time = 0
        leader = None
        tracking_stocks.clear()
        previous_rise_values.clear()
        leader_peak_rise = None
        leader_rise_before_decline = None
        first_condition_one_time = None
    else:
        message_log.append(
            (current_time_str,
             f"{RED}無法取得 {selected_symbol} 在 {current_time_str} 的價格，進場失敗{RESET}")
        )

def exit_trade(
    selected_stock_df, shares, entry_price, sell_cost,
    entry_fee, tax,
    message_log, current_time, hold_time, entry_time, use_f_exit=False
):
    global transaction_fee, transaction_discount, trading_tax
    global in_position, has_exited, current_position
    current_time_str = current_time if isinstance(current_time, str) else current_time.strftime('%H:%M:%S')
    selected_stock_df['time'] = pd.to_datetime(selected_stock_df['time'], format='%H:%M:%S').dt.time

    if isinstance(entry_time, str):
        entry_time_obj = datetime.strptime(entry_time, '%H:%M:%S').time()
    else:
        entry_time_obj = entry_time

    if use_f_exit:
        end_time = datetime.strptime('13:30', '%H:%M').time()
        end_price_series = selected_stock_df[selected_stock_df['time'] == end_time]['close']
        if not end_price_series.empty:
            end_price = end_price_series.values[0]
        else:
            print("無法取得 13:30 的數據，出場時間配對錯誤")
            message_log.append((current_time_str, f"{RED}出場時間配對錯誤{RESET}"))
            return None, None
        entry_datetime = datetime.combine(date.today(), entry_time_obj)
        if isinstance(current_time, datetime):
            current_datetime = current_time
        else:
            current_datetime = datetime.combine(date.today(), current_time)
        hold_time_calculated = int((current_datetime - entry_datetime).total_seconds() / 60)
    else:
        entry_index_series = selected_stock_df[selected_stock_df['time'] == entry_time_obj].index
        if not entry_index_series.empty:
            entry_index = entry_index_series[0]
            exit_index = entry_index + hold_time
            if exit_index >= len(selected_stock_df):
                print("出場時間超出範圍，無法進行交易")
                message_log.append((current_time_str, f"{RED}出場時間超出範圍{RESET}"))
                return None, None
            end_price = selected_stock_df.iloc[exit_index]['close']
        else:
            print("進場時間配對錯誤，無法找到精確的進場時間")
            message_log.append((current_time_str, f"{RED}進場時間配對錯誤{RESET}"))
            return None, None
        hold_time_calculated = hold_time

    buy_cost = shares * end_price * 1000
    exit_fee = int(buy_cost * (transaction_fee * 0.01) * (transaction_discount * 0.01))
    profit = sell_cost - buy_cost - entry_fee - exit_fee - tax
    return_rate = (profit * 100) / (buy_cost - exit_fee) if (buy_cost - exit_fee) != 0 else 0.0

    if use_f_exit:
        message_log.append(
            (current_time_str,
             f"{RED}股票出場，持有時間 {hold_time_calculated} 分鐘（強制出場）{RESET}")
        )
    else:
        message_log.append(
            (current_time_str,
             f"{RED}股票出場，持有時間 {hold_time_calculated} 分鐘{RESET}")
        )
    message_log.append(
        (current_time_str,
         f"{RED}持有張數：{shares} 張，出場價格：{end_price} 元，出場價金：{int(buy_cost)} 元，利潤：{int(profit)} 元，"
         f"報酬率：{return_rate:.2f}%，手續費：{exit_fee} 元{RESET}")
    )

    in_position = False
    has_exited = True
    return profit, return_rate

def consolidate_and_save_stock_symbols():
    mt_matrix_dict = load_mt_matrix_dict()
    matrix_dict_analysis = load_matrix_dict_analysis()
    
    if not mt_matrix_dict:
        print("mt_matrix_dict.json 文件不存在或為空，無法進行統整")
        return
    if not matrix_dict_analysis:
        print("matrix_dict_analysis.json 文件不存在或為空，無法進行統整")
        return
    consolidated_group_symbols = {group: [] for group in matrix_dict_analysis.keys()}
    
    for group, records in mt_matrix_dict.items():
        for record in records:
            if isinstance(record, dict):
                stock1 = record.get('stock1')
                stock2 = record.get('stock2')
                similarity_score = record.get('similarity_score', 0)
                
                if similarity_score >= 0.3:
                    for analysis_group, symbols in matrix_dict_analysis.items():
                        if stock1 in symbols and stock1 not in consolidated_group_symbols[analysis_group]:
                            consolidated_group_symbols[analysis_group].append(stock1)
                        if stock2 in symbols and stock2 not in consolidated_group_symbols[analysis_group]:
                            consolidated_group_symbols[analysis_group].append(stock2)
            else:
                print(f"警告：預期字典但獲得 {type(record)}，跳過該記錄。")
    
    for group in consolidated_group_symbols:
        consolidated_group_symbols[group] = list(set(consolidated_group_symbols[group]))
    nb_matrix_dict = {"consolidated_symbols": consolidated_group_symbols}
    save_nb_matrix_dict(nb_matrix_dict)
    print(f"統整後的股票代號已保存至 nb_matrix_dict.json，按族群分類。")

def calculate_kline_similarity(stock_data_list):
    similarity_results = []
    num_stocks = len(stock_data_list)
    for i in range(num_stocks):
        stock1 = stock_data_list[i]
        if 'symbol' not in stock1.columns:
            raise KeyError("DataFrame does not contain 'symbol' column.")
        symbol1 = stock1['symbol'].iloc[0]
        for j in range(i + 1, num_stocks):
            stock2 = stock_data_list[j]
            if 'symbol' not in stock2.columns:
                raise KeyError("DataFrame does not contain 'symbol' column.")
            symbol2 = stock2['symbol'].iloc[0]
            if symbol1 != symbol2:
                merged_df = pd.merge(stock1, stock2, on='time', suffixes=('_1', '_2'))
                merged_df['昨日收盤價_2'] = merged_df['昨日收盤價_2'].ffill().bfill()
                if 'high_1' not in merged_df.columns or 'high_2' not in merged_df.columns:
                    print(f"股票 {symbol1} 或 {symbol2} 缺少 'high' 欄位，跳過相似度計算。")
                    continue
                for col in ['open', 'high', 'low', 'close']:
                    merged_df[f'{col}_1_z'] = (merged_df[f'{col}_1'] - merged_df[f'{col}_1'].mean()) / merged_df[f'{col}_1'].std()
                    merged_df[f'{col}_2_z'] = (merged_df[f'{col}_2'] - merged_df[f'{col}_2'].mean()) / merged_df[f'{col}_2'].std()
                distance = np.sqrt(
                    (merged_df['open_1_z'] - merged_df['open_2_z']) ** 2 +
                    (merged_df['high_1_z'] - merged_df['high_2_z']) ** 2 +
                    (merged_df['low_1_z'] - merged_df['low_2_z']) ** 2 +
                    (merged_df['close_1_z'] - merged_df['close_2_z']) ** 2
                ).mean()
                similarity_score = 1 / (1 + distance)
                if similarity_score >= 0.3:
                    result = {
                        'stock1': symbol1,
                        'stock2': symbol2,
                        'similarity_score': similarity_score
                    }
                    similarity_results.append(result)
    if not similarity_results:
        print("沒有找到相似度大於等於 0.3 的結果")
        return pd.DataFrame(columns=['stock1', 'stock2', 'similarity_score'])
    similarity_df = pd.DataFrame(similarity_results)
    similarity_df = similarity_df.sort_values(by='similarity_score', ascending=False).reset_index(drop=True)
    return similarity_df

def calculate_limit_up_price(close_price):
    limit_up = close_price * 1.10
    if limit_up < 10:
        price_unit = 0.01
    elif limit_up < 50:
        price_unit = 0.05
    elif limit_up < 100:
        price_unit = 0.1
    elif limit_up < 500:
        price_unit = 0.5
    elif limit_up < 1000:
        price_unit = 1
    else:
        price_unit = 5
    limit_up_price = (limit_up // price_unit) * price_unit
    return limit_up_price

def save_mt_matrix_dict(mt_matrix_dict):
    with open('mt_matrix_dict.json', 'w', encoding='utf-8') as f:
        json.dump(mt_matrix_dict, f, indent=4, ensure_ascii=False, default=str)

def load_mt_matrix_dict():
    if os.path.exists('mt_matrix_dict.json'):
        with open('mt_matrix_dict.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        return {}

def load_nb_matrix_dict():
    if os.path.exists('nb_matrix_dict.json'):
        with open('nb_matrix_dict.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        return {}
    
def ensure_continuous_time_series(df):
    df['date'] = pd.to_datetime(df['date'])
    df['time'] = pd.to_datetime(df['time'], format='%H:%M:%S').dt.time

    full_time_index = pd.date_range(start='09:00', end='13:30', freq='1min').time
    full_index = pd.MultiIndex.from_product([df['date'].unique(), full_time_index], names=['date', 'time'])

    df.set_index(['date', 'time'], inplace=True)
    df = df.reindex(full_index)
    df[['symbol', '昨日收盤價', '漲停價']] = df[['symbol', '昨日收盤價', '漲停價']].ffill().bfill()

    if 'high' not in df.columns:
        df['high'] = df['close']
    if 'low' not in df.columns:
        df['low'] = df['close']

    df['close'] = df['close'].ffill()
    df['close'] = df['close'].fillna(df['昨日收盤價'])
    df['open'] = df['open'].ffill()
    df['open'] = df['open'].fillna(df['close'])
    df['high'] = df['high'].ffill()
    df['high'] = df['high'].fillna(df['close'])
    df['low'] = df['low'].ffill()
    df['low'] = df['low'].fillna(df['close'])
    df['volume'] = df['volume'].fillna(0)

    if '5min_pct_increase' not in df.columns:
        df['5min_pct_increase'] = 0.0
    else:
        df['5min_pct_increase'] = df['5min_pct_increase'].fillna(0.0)

    df.reset_index(inplace=True)
    return df
        
def load_disposition_stocks():
    disposition_file = 'Disposition.json'
    try:
        with open(disposition_file, 'r', encoding='utf-8') as f:
            disposition_data = json.load(f)
            return disposition_data
    except FileNotFoundError:
        print(f"錯誤：無法找到 {disposition_file} 文件。")
        return []
    except json.JSONDecodeError:
        print(f"錯誤：{disposition_file} 文件格式不正確。")
        return []
    
def fetch_disposition_stocks(client, matrix_dict_analysis):
    disposition_stocks = []
    for group, stock_list in matrix_dict_analysis.items():
        for symbol in stock_list:
            try:
                ticker_data = client.stock.intraday.ticker(symbol=symbol)
                if ticker_data.get('isDisposition', False):
                    disposition_stocks.append(symbol)
            except Exception as e:
                print(f"獲取 {symbol} 的處置股狀態時發生錯誤: {e}")
    with open('Disposition.json', 'w', encoding='utf-8') as f:
        json.dump(disposition_stocks, f, indent=4, ensure_ascii=False)

def calculate_average_over_high_list():
    while True:
        print('\n' + '=' * 50)
        print("選擇計算平均過高的模式：")
        print("1. 單一族群分析")
        print("2. 全部族群分析")
        print("0. 返回主選單")
        
        sub_choice = input("請輸入選項：")
        if sub_choice == '1':
            calculate_average_over_high()
        elif sub_choice == '2':
            matrix_dict_analysis = load_matrix_dict_analysis()
            all_group_names = list(matrix_dict_analysis.keys())
            if not all_group_names:
                print("沒有任何族群資料可供分析。")
                continue
            print("開始分析所有族群中的股票...")
            all_group_over_high_averages = []

            for i, group in enumerate(all_group_names):
                print(f"\n=== 分析族群：{group} ===")
                group_average = calculate_average_over_high(group_name=group)
                if group_average is not None:
                    all_group_over_high_averages.append(group_average)
                    
            if all_group_over_high_averages:
                overall_group_average = sum(all_group_over_high_averages) / len(all_group_over_high_averages)
                print(f"\n全部族群的平均過高間隔：{overall_group_average:.2f} 分鐘")
            else:
                print("\n沒有任何族群發生過高間隔的情形。")
        elif sub_choice == '0':
            main_menu()
        else:
            print("無效的選項，請重新輸入")

def load_kline_data():
    daily_kline_data = {}
    intraday_kline_data = {}

    if os.path.exists('daily_kline_data.json'):
        with open('daily_kline_data.json', 'r', encoding='utf-8') as f:
            try:
                daily_kline_data = json.load(f)
                if not daily_kline_data:
                    print("日K線數據檔案為空，請先更新數據。")
            except json.JSONDecodeError:
                print("日K線數據檔案格式錯誤，請先更新數據。")

    if os.path.exists('intraday_kline_data.json'):
        with open('intraday_kline_data.json', 'r', encoding='utf-8') as f:
            try:
                intraday_kline_data = json.load(f)
                if not intraday_kline_data:
                    print("一分K線數據檔案為空，請先更新數據。")
            except json.JSONDecodeError:
                print("一分K線數據檔案格式錯誤，請先更新數據。")

    return daily_kline_data, intraday_kline_data

def calculate_average_over_high(group_name=None):
    daily_kline_data, intraday_kline_data = load_kline_data()

    matrix_dict_analysis = load_matrix_dict_analysis()
    
    if group_name is None:
        group_name = input("請輸入要分析的族群名稱：")
    
    if group_name not in matrix_dict_analysis:
        print("沒有此族群資料")
        return None

    symbols_to_analyze = matrix_dict_analysis[group_name]
    disposition_stocks = load_disposition_stocks()
    symbols_to_analyze = [symbol for symbol in symbols_to_analyze if symbol not in disposition_stocks]

    if not symbols_to_analyze:
        print(f"{group_name} 中沒有可供分析的股票。")
        return None

    print(f"開始分析族群 {group_name} 中的股票...")
    any_condition_one_triggered = False 
    group_over_high_averages = []

    for symbol in symbols_to_analyze:
        print(f"\n正在分析股票：{symbol}")
        
        if symbol not in daily_kline_data or symbol not in intraday_kline_data:
            print(f"無法取得 {symbol} 的日 K 線或一分 K 線數據，跳過。")
            continue
        
        daily_kline_df = pd.DataFrame(daily_kline_data[symbol])
        intraday_data = pd.DataFrame(intraday_kline_data[symbol])

        condition_one_triggered = False
        condition_two_triggered = False
        previous_high = None
        condition_two_time = None
        over_high_intervals = []

        for idx, row in intraday_data.iterrows():
            current_time = pd.to_datetime(row['time']).time()
            if previous_high is None:
                previous_high = row['high']
                continue

            if not condition_one_triggered:
                if row['5min_pct_increase'] >= 2:
                    condition_one_triggered = True
                    condition_two_triggered = False
                    any_condition_one_triggered = True

                    print(f"{symbol} 觸發條件一，開始監測五分鐘漲幅，五分鐘漲幅: {row['5min_pct_increase']:.2f}%")

            if condition_one_triggered and not condition_two_triggered:
                if row['high'] <= previous_high:
                    current_time_str = current_time.strftime('%H:%M:%S')
                    print(f"{symbol} 觸發條件二！時間：{current_time_str}")

                    condition_two_time = current_time
                    condition_two_triggered = True

            elif condition_two_triggered:
                if row['highest'] > previous_high:
                    condition_three_time_str = current_time.strftime('%H:%M:%S')
                    print(f"{symbol} 觸發條件三！時間：{condition_three_time_str}")
                    if condition_two_time:
                        today = datetime.today().date()
                        condition_two_datetime = datetime.combine(today, condition_two_time)
                        condition_three_datetime = datetime.combine(today, current_time)
                        interval = (condition_three_datetime - condition_two_datetime).total_seconds() / 60
                        print(f"{symbol} 過高間隔：{interval:.2f} 分鐘")
                        over_high_intervals.append(interval)

                    condition_one_triggered = False
                    condition_two_triggered = False
                    condition_two_time = None

            previous_high = row['high']

        if over_high_intervals:
            q1 = np.percentile(over_high_intervals, 25)
            q3 = np.percentile(over_high_intervals, 75)
            iqr = q3 - q1
            lower_bound = q1 - 1.5 * iqr
            upper_bound = q3 + 1.5 * iqr
            filtered_intervals = [interval for interval in over_high_intervals if lower_bound <= interval <= upper_bound]
            if filtered_intervals:
                average_interval = sum(filtered_intervals) / len(filtered_intervals)
                print(f"{symbol} 平均過高間隔：{average_interval:.2f} 分鐘")
                group_over_high_averages.append(average_interval)
            else:
                print(f"{symbol} 沒有有效的過高間隔數據")
        else:
            print(f"{symbol} 沒有觸發過高間隔的情形")

    if group_over_high_averages:
        group_average_over_high = sum(group_over_high_averages) / len(group_over_high_averages)
        print(f"{group_name} 平均過高間隔：{group_average_over_high:.2f} 分鐘")
        return group_average_over_high
    else:
        print(f"{group_name} 沒有有效的過高間隔數據")
        return None

def main_menu():
    global capital_per_stock
    load_settings()
    print('\n' + '=' * 50)
    print(f"\n目前股票的單筆投入資本額為{capital_per_stock}萬元")
    while True:
        print("請選擇功能：")
        print("1. 回測模式")
        print("2. 模擬交易模式")
        print("3. 正式下單模式")
        print("4. 管理族群")
        print("5. 設定選單")
        print("6. 更新K線數據")
        print("7. 查詢處置股清單")
        print("0. 退出程式")
        print('\n' + '=' * 50)
        choice = input("請輸入選項：")
        if choice == '1':
            backtesting_menu_list()
        elif choice == '2':
            simulate_trading_menu_list()
        elif choice == '3':
            print('目前仍在測試中，敬請期待')
            main_menu()
        elif choice == '4':
            manage_groups()
        elif choice == '5':
            settings_menu()
        elif choice == '6':
            update_kline_data_menu()
        elif choice == '7':
            display_disposition_stocks()
        elif choice == '0':
            print("退出程式...下次再見")
            break
        else:
            print("無效的選項，請重新輸入")

def backtesting_menu_list():
    print('\n' + '=' * 50)
    print("\n請選擇功能：")
    print("1. 計算平均過高、2. 自選進場模式、3. 極大化利潤模式、0. 返回主選單")
    print('\n' + '=' * 50)
    back_choice = input("請選擇功能：")
    if back_choice == '1':
        calculate_average_over_high_list()
    elif back_choice == '2':
        simulate_trading_menu()
    elif back_choice == '3':
        maximize_profit_analysis()
    elif back_choice == '0':
        main_menu()
    else:
        print("無效的選項，請重新輸入")

def simulate_trading_menu_list():
    print('\n' + '=' * 50)
    print("\n請選擇功能：")
    print("1. 開始模擬交易、2. 登入帳戶、3. 修改api金鑰、0. 返回主選單")
    print('\n' + '=' * 50)
    back_choice = input("請選擇功能：")
    if back_choice == '1':
        start_trading()
    elif back_choice == '2':
        login()
    elif back_choice == '3':
        print('目前仍在測試中，敬請期待')
        simulate_trading_menu_list()
    elif back_choice == '0':
        main_menu()
    else:
        print("無效的選項，請重新輸入")

capital_per_stock = 0
transaction_fee = 0
transaction_discount = 0
trading_tax = 0
below_50 = 0
price_gap_50_to_100 = 0
price_gap_100_to_500 = 0
price_gap_500_to_1000 = 0
price_gap_above_1000 = 0
allow_reentry_after_stop_loss = False

def load_symbols_to_analyze():
    nb_matrix_dict = load_nb_matrix_dict()
    consolidated_symbols = nb_matrix_dict.get("consolidated_symbols", {})
    symbols = []
    for group_symbols in consolidated_symbols.values():
        symbols.extend(group_symbols)
    disposition_stocks = load_disposition_stocks()
    symbols = [symbol for symbol in symbols if symbol not in disposition_stocks]
    return symbols

def load_group_symbols():
    if not os.path.exists('nb_matrix_dict.json'):
        print("nb_matrix_dict.json 文件不存在。")
        return {}
    with open('nb_matrix_dict.json', 'r', encoding='utf-8') as f:
        group_symbols = json.load(f)
    return group_symbols

def start_trading():
    client, api_key = init_fugle_client()
    symbols_to_analyze = load_symbols_to_analyze()
    stop_trading = False
    max_symbols_to_fetch = 20

    group_symbols = load_group_symbols()

    if not group_symbols:
        print("沒有加載到任何族群資料，請確認 nb_matrix_dict.json 的存在與內容。")
        return  

    consolidated_symbols = group_symbols.get('consolidated_symbols', {})
    if not consolidated_symbols:
        print("沒有找到 'consolidated_symbols'，請確認資料結構。")
        return

    group_positions = {group: False for group in consolidated_symbols.keys()}
    try:
        wait_minutes = int(input("請輸入等待時間（分鐘）："))
    except ValueError:
        print("等待時間必須是整數。")
        return

    hold_minutes_input = input("請輸入持有時間（分鐘，輸入 'F' 代表持有到13:30強制出場）：")
    if hold_minutes_input.upper() == 'F':
        hold_minutes = None
    else:
        try:
            hold_minutes = int(hold_minutes_input)
        except ValueError:
            print("持有時間必須是整數或 'F'。")
            return

    existing_auto_daily_data = {}
    if os.path.exists('auto_daily.json'):
        with open('auto_daily.json', 'r', encoding='utf-8') as f:
            try:
                existing_auto_daily_data = json.load(f)
            except json.JSONDecodeError:
                existing_auto_daily_data = {}
    else:
        print("auto_daily.json 不存在，將建立新的。")

    print("開始取得日K線數據並與現有資料比對...")
    auto_daily_data = {}
    data_is_same = True
    initial_api_count = 0
    symbols_fetched = 0

    for symbol in symbols_to_analyze[:max_symbols_to_fetch]:
        if initial_api_count >= 55:
            print("已達到55次API請求，休息1分鐘...")
            time_module.sleep(60)
            initial_api_count = 0
        daily_kline_df = fetch_daily_kline_data(client, symbol, days=2)
        initial_api_count += 1
        if daily_kline_df.empty:
            print(f"無法取得 {symbol} 的日K數據，跳過。")
            continue
        daily_kline_data = daily_kline_df.to_dict(orient='records')
        auto_daily_data[symbol] = daily_kline_data
        existing_data = existing_auto_daily_data.get(symbol)
        if existing_data != daily_kline_data:
            data_is_same = False
            print(f"{symbol} 的日K數據與現有資料不同，將更新資料。")
            existing_auto_daily_data[symbol] = daily_kline_data
        else:
            print(f"{symbol} 的日K數據與現有資料相同，跳過更新。")
        symbols_fetched += 1

    if not data_is_same:
        remaining_symbols = symbols_to_analyze[max_symbols_to_fetch:]
        print(f"發現前 {max_symbols_to_fetch} 支股票的日K數據有更新，開始取得剩餘股票的日K數據並更新。")
        for symbol in remaining_symbols:
            if initial_api_count >= 55:
                print("已達到55次API請求，休息1分鐘...")
                time_module.sleep(60)
                initial_api_count = 0
            daily_kline_df = fetch_daily_kline_data(client, symbol, days=2)
            initial_api_count += 1
            if daily_kline_df.empty:
                print(f"無法取得 {symbol} 的日K數據，跳過。")
                continue
            daily_kline_data = daily_kline_df.to_dict(orient='records')
            auto_daily_data[symbol] = daily_kline_data
            existing_data = existing_auto_daily_data.get(symbol)
            if existing_data != daily_kline_data:
                print(f"{symbol} 的日K數據與現有資料不同，將更新資料。")
                existing_auto_daily_data[symbol] = daily_kline_data
            else:
                print(f"{symbol} 的日K數據與現有資料相同，跳過更新。")

    if symbols_fetched < max_symbols_to_fetch:
        print(f"注意：僅取得了 {symbols_fetched} 支股票的日K數據。")

    with open('auto_daily.json', 'w', encoding='utf-8') as f:
        json.dump(existing_auto_daily_data, f, ensure_ascii=False, indent=4)
    print("已更新 auto_daily.json。")

    print("開始補齊一分K數據。")

    trading_day = get_recent_trading_day().strftime('%Y-%m-%d')
    yesterday_close_prices = {}
    for symbol in symbols_to_analyze:
        daily_data = existing_auto_daily_data.get(symbol, [])
        if not daily_data:
            print(f"無法從 auto_daily.json 獲取 {symbol} 的日K數據，將使用最新取得的數據。")
            print(f"正在嘗試為 {symbol} 獲取日K數據...")
            daily_kline_df = fetch_daily_kline_data(client, symbol, days=5)
            if not daily_kline_df.empty:
                daily_kline_data = daily_kline_df.to_dict(orient='records')
                existing_auto_daily_data[symbol] = daily_kline_data
                with open('auto_daily.json', 'w', encoding='utf-8') as f:
                    json.dump(existing_auto_daily_data, f, ensure_ascii=False, indent=4)
                print(f"已成功為 {symbol} 補上日K數據。")
                if len(daily_kline_data) > 1:
                    # 控制昨日收盤價取得的項數
                    yesterday_close = daily_kline_data[0].get('close', 0)
                    yesterday_close_prices[symbol] = yesterday_close
                else:
                    print(f"警告：{symbol} 的日K數據數量不足，無法獲取昨日收盤價。")
                    yesterday_close_prices[symbol] = 0
            else:
                print(f"警告：{symbol} 沒有任何日K線數據。")
                yesterday_close_prices[symbol] = 0
        else:
            sorted_daily_data = sorted(daily_data, key=lambda x: x['date'], reverse=True)
            latest_trading_day_str = sorted_daily_data[0]['date']
            latest_trading_day = datetime.strptime(latest_trading_day_str, '%Y-%m-%d')
            expected_trading_day = datetime.strptime(trading_day, '%Y-%m-%d')
            if (expected_trading_day - latest_trading_day).days > 1:
                print(f"警告：{symbol} 的最新交易日 {latest_trading_day.strftime('%Y-%m-%d')} 與預期交易日 {trading_day} 不符。")
            if len(sorted_daily_data) > 1:
                # 控制昨日收盤價取得的項數
                yesterday_close = sorted_daily_data[0].get('close', 0)
                yesterday_close_prices[symbol] = yesterday_close
            else:
                print(f"警告：{symbol} 沒有上一個交易日的收盤價資料。")
                yesterday_close_prices[symbol] = 0

    current_time = datetime.now()
    market_end_time = current_time.replace(hour=13, minute=30, second=0, microsecond=0)

    pre_market_start = current_time.replace(hour=0, minute=0, second=0, microsecond=0)
    pre_market_end = current_time.replace(hour=8, minute=30, second=0, microsecond=0)

    if pre_market_start <= current_time < pre_market_end:
        trading_day = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        initial_fetch_end_time_str = "13:30"
        print(f"目前時間在 00:00 ~ 08:30，將取得前一日 {trading_day} 的一分K數據。")
    elif current_time > market_end_time:
        initial_fetch_end_time = market_end_time
        initial_fetch_end_time_str = initial_fetch_end_time.strftime('%H:%M')
    else:
        initial_fetch_end_time = (current_time - timedelta(minutes=1)).replace(second=0, microsecond=0)
        initial_fetch_end_time_str = initial_fetch_end_time.strftime('%H:%M')

    auto_intraday_data = {}
    initial_api_count = 0

    with ThreadPoolExecutor(max_workers=20) as executor:
        future_to_symbol = {}
        for symbol in symbols_to_analyze:
            if initial_api_count >= 200:
                print("已達到200次API請求，休息1分鐘...")
                time_module.sleep(60)
                initial_api_count = 0
            if pre_market_start <= current_time < pre_market_end:
                print(f"正在取得 {symbol} 的一分K數據從 09:00 到 {initial_fetch_end_time_str} (前一日)...")
                trading_day_to_fetch = trading_day
            else:
                print(f"正在取得 {symbol} 的一分K數據從 09:00 到 {initial_fetch_end_time_str}...")
                trading_day_to_fetch = trading_day
            yesterday_close = yesterday_close_prices.get(symbol, 0)
            if yesterday_close == 0:
                print(f"警告：{symbol} 的昨日收盤價為0，將跳過一分K數據的獲取。")
                continue
            future = executor.submit(
                fetch_intraday_data,
                client=client,
                symbol=symbol,
                trading_day=trading_day_to_fetch,
                yesterday_close_price=yesterday_close,
                start_time="09:00",
                end_time=initial_fetch_end_time_str
            )
            future_to_symbol[future] = symbol
            initial_api_count += 1

        for future in as_completed(future_to_symbol):
            symbol = future_to_symbol[future]
            try:
                intraday_df = future.result()
                if intraday_df.empty:
                    print(f"無法取得 {symbol} 的一分K數據，跳過。")
                    continue
                intraday_data = intraday_df.to_dict(orient='records')
                intraday_data_sorted = sorted(intraday_data, key=lambda x: x['time'])
                existing_candles = []
                calculated_candles = []
                for candle in intraday_data_sorted:
                    calculated_candle = calculate_5min_pct_increase(candle, existing_candles)
                    if '漲停價' in calculated_candle:
                        calculated_candle['漲停價'] = truncate_to_two_decimals(calculated_candle['漲停價'])
                    calculated_candles.append(calculated_candle)
                    existing_candles.append(calculated_candle)
                auto_intraday_data[symbol] = calculated_candles
            except Exception as e:
                print(f"在獲取 {symbol} 的一分K數據時發生錯誤：{e}")

    save_auto_intraday_data(auto_intraday_data)
    print("已更新 auto_intraday.json。")

    current_time = datetime.now()
    current_time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')

    pre_market_start = current_time.replace(hour=8, minute=30, second=0, microsecond=0)
    pre_market_end = current_time.replace(hour=8, minute=59, second=59, microsecond=59)
    market_start = current_time.replace(hour=9, minute=0, second=0, microsecond=0)
    market_end = current_time.replace(hour=23, minute=30, second=0, microsecond=0)

    if pre_market_start <= current_time < pre_market_end:
        print(f"目前為 {current_time_str}，盤前準備時間。")
    elif market_start <= current_time <= market_end:
        print(f"目前為 {current_time_str}，盤中交易時間。")
        print("開始監控，即時取得一分K數據。")
        print("輸入 'Q' 返回主選單：", end='', flush=True)

        group_position = False
        has_exited = False
        current_position = None
        hold_time = 0
        message_log = []
        already_entered_stocks = []
        stop_loss_triggered = False
        final_check_active = False
        final_check_count = 0
        in_waiting_period = False
        waiting_time = 0
        leader = None
        tracking_stocks = set()
        previous_rise_values = {}
        leader_peak_rise = None
        leader_rise_before_decline = None
        first_condition_one_time = None
        can_trade = True

        while not stop_trading:
            current_time = datetime.now()
            if market_start <= current_time <= market_end:
                wait_until_next_minute()
                fetch_time = datetime.now() - timedelta(minutes=1)
                trading_day = fetch_time.strftime('%Y-%m-%d')
                fetch_time_str = fetch_time.strftime('%H:%M')

                if fetch_time.time() > market_end.time():
                    fetch_time_str = "13:30"

                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                print("\n" + "=" * 50)
                print(f"\n{timestamp} 市場開盤中，取得 {fetch_time_str} 分的即時一分K數據。")
                if market_start <= current_time <= market_end:
                    current_time_str = current_time.strftime('%H:%M')
                    print(f"正在取得一分K數據從 {current_time_str} 到 {current_time_str}...")
                else:
                    print("正在取得一分K數據從 09:00 到 13:30...")

                updated_intraday_data = {}
                with ThreadPoolExecutor(max_workers=20) as executor:
                    future_to_symbol = {}
                    for symbol in symbols_to_analyze:
                        yesterday_close = yesterday_close_prices.get(symbol, 1)
                        if yesterday_close == 0:
                            continue
                        future = executor.submit(
                            fetch_intraday_data,
                            client=client,
                            symbol=symbol,
                            trading_day=trading_day,
                            yesterday_close_price=yesterday_close,
                            start_time=fetch_time_str,
                            end_time=fetch_time_str
                        )
                        future_to_symbol[future] = symbol

                    for future in as_completed(future_to_symbol):
                        symbol = future_to_symbol[future]
                        try:
                            intraday_df = future.result()
                            if intraday_df.empty:
                                print(f"無法取得 {symbol} 的一分K數據，跳過。")
                                continue
                            latest_candle = intraday_df.to_dict(orient='records')[0]
                            if not all(key in latest_candle and latest_candle[key] is not None for key in ['open', 'close', 'high', 'low']):
                                print(f"{symbol} 的最新一分K數據不完整，跳過該數據。")
                                continue
                            latest_candle = calculate_5min_pct_increase(latest_candle, auto_intraday_data.get(symbol, []))
                            if '漲停價' in latest_candle:
                                latest_candle['漲停價'] = truncate_to_two_decimals(latest_candle['漲停價'])
                            if latest_candle.get('volume', 0) == 0:
                                latest_candle['low'] = latest_candle['open']
                                latest_candle['high'] = latest_candle['open']
                                latest_candle['close'] = latest_candle['open']
                            if symbol not in updated_intraday_data:
                                updated_intraday_data[symbol] = []
                            updated_intraday_data[symbol].append(latest_candle)
                        except Exception as e:
                            print(f"在獲取 {symbol} 的一分K數據時發生錯誤：{e}")

                for symbol, candles in updated_intraday_data.items():
                    if symbol not in auto_intraday_data:
                        auto_intraday_data[symbol] = []
                    auto_intraday_data[symbol].extend(candles)
                    auto_intraday_data[symbol] = auto_intraday_data[symbol][-1000:]

                save_auto_intraday_data(auto_intraday_data)
                print("一分K數據已成功處理並返回。")
                print("=" * 50)

                process_live_trading_logic(
                    symbols_to_analyze,
                    fetch_time_str,
                    wait_minutes,
                    hold_minutes,
                    message_log,
                    False,
                    has_exited,
                    current_position,
                    hold_time,
                    already_entered_stocks,
                    stop_loss_triggered,
                    final_check_active,
                    final_check_count,
                    in_waiting_period,
                    waiting_time,
                    leader,
                    tracking_stocks,
                    previous_rise_values,
                    leader_peak_rise,
                    leader_rise_before_decline,
                    first_condition_one_time,
                    can_trade,
                    group_positions
                )
            else:
                if pre_market_start <= current_time < pre_market_end:
                    print(f"目前為 {current_time.strftime('%Y-%m-%d %H:%M:%S')}，盤前時間。")
                    print("開始更新前一日的一分K數據。")
                    end_time_str = "13:30"
                    updated_intraday_data = {}
                    with ThreadPoolExecutor(max_workers=20) as executor:
                        future_to_symbol = {}
                        for symbol in symbols_to_analyze:
                            yesterday_close = yesterday_close_prices.get(symbol, 0)
                            if yesterday_close == 0:
                                continue
                            future = executor.submit(
                                fetch_intraday_data,
                                client=client,
                                symbol=symbol,
                                trading_day=trading_day,
                                yesterday_close_price=yesterday_close,
                                start_time="09:00",
                                end_time=end_time_str
                            )
                            future_to_symbol[future] = symbol

                        for future in as_completed(future_to_symbol):
                            symbol = future_to_symbol[future]
                            try:
                                intraday_df = future.result()
                                if intraday_df.empty:
                                    print(f"無法取得 {symbol} 的一分K數據，跳過。")
                                    continue
                                intraday_data = intraday_df.to_dict(orient='records')
                                intraday_data_sorted = sorted(intraday_data, key=lambda x: x['time'])
                                existing_candles = []
                                calculated_candles = []
                                for candle in intraday_data_sorted:
                                    calculated_candle = calculate_5min_pct_increase(candle, existing_candles)
                                    if '漲停價' in calculated_candle:
                                        calculated_candle['漲停價'] = truncate_to_two_decimals(calculated_candle['漲停價'])
                                    calculated_candles.append(calculated_candle)
                                    existing_candles.append(calculated_candle)
                                if symbol not in updated_intraday_data:
                                    updated_intraday_data[symbol] = []
                                updated_intraday_data[symbol].extend(calculated_candles)
                            except Exception as e:
                                print(f"在獲取 {symbol} 的一分K數據時發生錯誤：{e}")

                    for symbol, candles in updated_intraday_data.items():
                        if symbol not in auto_intraday_data:
                            auto_intraday_data[symbol] = []
                        auto_intraday_data[symbol].extend(candles)
                        auto_intraday_data[symbol] = auto_intraday_data[symbol][-1000:]

                    save_auto_intraday_data(auto_intraday_data)
                    print("一分K數據已成功處理並返回。")
                    print("已更新 auto_intraday.json。")
                    print(f"目前為 {current_time_str}，盤前時間。")
                elif current_time > market_end:
                    print(f"目前為 {current_time.strftime('%Y-%m-%d %H:%M:%S')}，盤後時間。")
                    print("開始更新一分K數據。")
                    if current_time.time() > market_end.time():
                        end_time_str = "13:30"
                    else:
                        end_time_str = current_time.strftime('%H:%M')
                    print(f"正在取得一分K數據從 09:00 到 {end_time_str}...")
                    updated_intraday_data = {}
                    with ThreadPoolExecutor(max_workers=20) as executor:
                        future_to_symbol = {}
                        for symbol in symbols_to_analyze:
                            yesterday_close = yesterday_close_prices.get(symbol, 0)
                            if yesterday_close == 0:
                                continue
                            future = executor.submit(
                                fetch_intraday_data,
                                client=client,
                                symbol=symbol,
                                trading_day=trading_day,
                                yesterday_close_price=yesterday_close,
                                start_time="09:00",
                                end_time=end_time_str
                            )
                            future_to_symbol[future] = symbol

                        for future in as_completed(future_to_symbol):
                            symbol = future_to_symbol[future]
                            try:
                                intraday_df = future.result()
                                if intraday_df.empty:
                                    print(f"無法取得 {symbol} 的一分K數據，跳過。")
                                    continue
                                intraday_data = intraday_df.to_dict(orient='records')
                                intraday_data_sorted = sorted(intraday_data, key=lambda x: x['time'])
                                existing_candles = []
                                calculated_candles = []
                                for candle in intraday_data_sorted:
                                    calculated_candle = calculate_5min_pct_increase(candle, existing_candles)
                                    if '漲停價' in calculated_candle:
                                        calculated_candle['漲停價'] = truncate_to_two_decimals(calculated_candle['漲停價'])
                                    calculated_candles.append(calculated_candle)
                                    existing_candles.append(calculated_candle)
                                if symbol not in updated_intraday_data:
                                    updated_intraday_data[symbol] = []
                                updated_intraday_data[symbol].extend(calculated_candles)
                            except Exception as e:
                                print(f"在獲取 {symbol} 的一分K數據時發生錯誤：{e}")

                    for symbol, candles in updated_intraday_data.items():
                        if symbol not in auto_intraday_data:
                            auto_intraday_data[symbol] = []
                        auto_intraday_data[symbol].extend(candles)
                        auto_intraday_data[symbol] = auto_intraday_data[symbol][-1000:]

                    save_auto_intraday_data(auto_intraday_data)
                    print("一分K數據已成功處理並返回。")
                    print("已更新 auto_intraday.json。")
                    print(f"目前為 {current_time_str}，盤後時間。")
                else:
                    print(f"目前為 {current_time.strftime('%Y-%m-%d %H:%M:%S')}，非盤前、盤中、盤後時間。")

            if user_wants_to_quit():
                print("\n收到退出指令，停止交易...")
                stop_trading = True

        print("已停止交易，返回主選單")
    else:
        print("目前非交易時間，已補齊最近交易日的一分K數據，返回主選單。")

def login():
    file_path = "shioaji_logic.py"  # 要更新的檔案路徑

    print('\n' + '=' * 50 + '\n')
    print("當前 api key 為：" + shioaji_logic.TEST_API_KEY)
    print("當前憑證路徑為：" + shioaji_logic.CA_CERT_PATH)
    print("當前憑證密碼為：" + shioaji_logic.CA_PASSWORD)
    print('\n' + '=' * 50)
    print("1. 修改 api key、2. 修改 api secret、3. 修改憑證路徑、4. 修改憑證密碼")
    api_setting = input("請選擇功能：")
    if api_setting == "1":
        new_api_key = input("請輸入新的 api key：")
        update_variable(file_path, "TEST_API_KEY", new_api_key)
    elif api_setting == "2":
        new_api_secret = input("請輸入新的 api secret：")
        update_variable(file_path, "TEST_API_SECRET", new_api_secret)
    elif api_setting == "3":
        new_ca_path = input("請輸入新的憑證路徑：")
        update_variable(file_path, "CA_CERT_PATH", new_ca_path, is_raw=True)
    elif api_setting == "4":
        new_ca_password = input("請輸入新的憑證密碼：")
        update_variable(file_path, "CA_PASSWORD", new_ca_password)
    else:
        print("請輸入合法字元...")
        login()

def update_variable(file_path, var_name, new_value, is_raw=False):
    """
    更新指定檔案中以 var_name 開頭的變數的值。
    若 is_raw 為 True，則會以 raw 字串格式儲存（例如 CA_CERT_PATH）
    """
    lines = []
    with open(file_path, "r", encoding="utf-8") as f:
        for line in f:
            # 如果該行以變數名稱開頭，則替換該行
            if line.lstrip().startswith(var_name + " ="):
                if is_raw:
                    new_line = f'{var_name} = r"{new_value}"\n'
                else:
                    new_line = f'{var_name} = "{new_value}"\n'
                lines.append(new_line)
            else:
                lines.append(line)
    with open(file_path, "w", encoding="utf-8") as f:
        f.writelines(lines)
    print(f"{var_name} 已更新為: {new_value}")
    importlib.reload(shioaji_logic)

#登入
api = sj.Shioaji(simulation=True)
accounts = api.login(api_key = shioaji_logic.TEST_API_KEY, secret_key = shioaji_logic.TEST_API_SECRET)
api.activate_ca(
    ca_path=shioaji_logic.CA_CERT_PATH,
    ca_passwd=shioaji_logic.CA_PASSWORD
)
'''
print("ca_path:", shioaji_logic.CA_CERT_PATH)
print("ca_password:", shioaji_logic.CA_PASSWORD)
'''
#新增管理套件
to = tp.TouchOrderExecutor(api)

def process_live_trading_logic(
    symbols_to_analyze,
    current_time_str,
    wait_minutes,
    hold_minutes,
    message_log,
    in_position,
    has_exited,
    current_position,
    hold_time,
    already_entered_stocks,
    stop_loss_triggered,
    final_check_active,
    final_check_count,
    in_waiting_period,
    waiting_time,
    leader,
    tracking_stocks,
    previous_rise_values,
    leader_peak_rise,
    leader_rise_before_decline,
    first_condition_one_time,
    can_trade,
    group_positions,
    nb_matrix_path='nb_matrix_dict.json'
):
    """
    盤中進場邏輯（進場條件分為【漲停進場】與【拉高進場】）：
      1. 檢查每支股票是否觸發進場條件：
           - 若最新K線的 high 等於漲停價，且前一根K線的 high 小於漲停價，則觸發【漲停進場】；
           - 若 5min_pct_increase >= 2.0 並且當前K棒成交量大於1.5倍開盤前三分鐘（09:00-09:02）的平均成交量，則觸發【拉高進場】。
      2. 將觸發股票所屬族群標記為「觀察中」，記錄觸發方式與開始等待時間。
      3. 若族群觸發【漲停進場】，每分鐘更新該族群追蹤清單並廣播等待狀態，
         待等待時間滿後從追蹤清單中選擇中間偏後的股票進場，並以全新格式廣播進場結果，
         同時計算進場後的停損方式與停損價。
      4. 若族群觸發【拉高進場】，則：
           - 首次選定領漲股票時廣播（例如：「拉高進場 化學 族群，領漲：1711」）；
           - 當領漲股票反轉（即當前K線的 high <= 前一K線的 high）時廣播反轉並開始等待（例如：「拉高進場 化學 族群，領漲 1711 反轉，開始等待。」）；  
             此時記錄等待起始時間與從該分鐘開始的等待計數；
           - 在等待期間，每分鐘以 print 輸出等待狀態訊息；
           - 在等待期間，如追蹤清單中有其他股票的漲幅超過當前記錄的「領漲反轉漲幅」，則進行【領漲替換】並廣播替換訊息；
           - 待等待時間累計達到設定值後，從追蹤清單中排除領漲股票後，依排序選出中間偏後的股票進場，
             並以全新格式廣播進場結果，同時計算停損方式與停損價。
      5. 進場後，將族群狀態更新為「已進場」。
    """
    # 引入全域參數
    global capital_per_stock, transaction_fee, transaction_discount, trading_tax
    global price_gap_below_50, price_gap_50_to_100, price_gap_100_to_500, price_gap_500_to_1000, price_gap_above_1000

    try:
        current_time = datetime.strptime(current_time_str, '%H:%M')
    except ValueError:
        print(f"時間格式錯誤：{current_time_str}，應為 'HH:MM' 格式。")
        return
    trading_time = current_time.time()
    trading_time_str = current_time.strftime('%H:%M:%S')
    print(f"當前交易時間：{trading_time_str}")

    # 讀取族群資料
    if os.path.exists(nb_matrix_path):
        with open(nb_matrix_path, 'r', encoding='utf-8') as f:
            try:
                nb_matrix_dict = json.load(f)
                if not isinstance(nb_matrix_dict, dict):
                    print(f"{nb_matrix_path} 的格式不正確，預期為字典。")
                    return
            except json.JSONDecodeError:
                print(f"{nb_matrix_path} 的格式不正確，無法解析。")
                return
    else:
        print(f"無法找到 {nb_matrix_path}，無法提取股票代號。")
        return

    if "consolidated_symbols" not in nb_matrix_dict:
        print(f"{nb_matrix_path} 中缺少 'consolidated_symbols' 鍵。")
        return

    consolidated_symbols = nb_matrix_dict["consolidated_symbols"]
    if not isinstance(consolidated_symbols, dict):
        print(f"'consolidated_symbols' 的格式不正確，預期為字典。")
        return

    # 讀取 auto_intraday.json 的即時一分K數據
    auto_intraday_path = 'auto_intraday.json'
    if os.path.exists(auto_intraday_path):
        with open(auto_intraday_path, 'r', encoding='utf-8') as f:
            try:
                auto_intraday_data = json.load(f)
                if not isinstance(auto_intraday_data, dict):
                    print(f"{auto_intraday_path} 的格式不正確，預期為字典。")
                    return
            except json.JSONDecodeError:
                print(f"{auto_intraday_path} 的格式不正確，無法解析。")
                return
    else:
        print(f"無法找到 {auto_intraday_path}，無法進行交易判斷。")
        return

    # 建立每支股票的 DataFrame
    stock_data_collection = {}
    for sym in symbols_to_analyze:
        if sym in auto_intraday_data:
            df = pd.DataFrame(auto_intraday_data[sym])
            try:
                df['time'] = pd.to_datetime(df['time'], format='%H:%M:%S').dt.time
            except Exception as e:
                print(f"轉換 'time' 欄位時出錯：{e}")
                stock_data_collection[sym] = pd.DataFrame()
                continue
            df = df.sort_values(by='time').reset_index(drop=True)
            stock_data_collection[sym] = df
        else:
            stock_data_collection[sym] = pd.DataFrame()

    # 1. 檢查各股票是否觸發進場條件
    eligible_stocks_for_entry = []
    unpositioned_groups = [group for group, status in group_positions.items() if not status or status == False]
    for group in unpositioned_groups:
        if group not in consolidated_symbols:
            print(f"警告：族群 {group} 不存在於 consolidated_symbols 中，跳過。")
            continue
        stock_symbols = consolidated_symbols[group]
        if not stock_symbols:
            print(f"警告：族群 {group} 沒有股票代號列表。")
            continue
        for stock_symbol in stock_symbols:
            if stock_symbol.strip().upper() not in [s.strip().upper() for s in symbols_to_analyze]:
                print(f"股票 {stock_symbol} 不在分析列表中，跳過。")
                continue
            df = stock_data_collection.get(stock_symbol)
            if df is None or df.empty:
                print(f"股票 {stock_symbol} 沒有K線數據，跳過。")
                continue
            current_rows = df[df['time'] == trading_time]
            if current_rows.empty:
                print(f"股票 {stock_symbol} 在 {trading_time_str} 沒有K線數據，跳過。")
                continue
            row = current_rows.iloc[0]
            high = row.get('high', 0.0)
            limit_up_price = row.get('漲停價', 0.0)
            five_min_increase = row.get('5min_pct_increase', 0.0)
            prev_time = (datetime.combine(date.today(), trading_time) - timedelta(minutes=1)).time()
            prev_rows = df[df['time'] == prev_time]
            previous_high = prev_rows.iloc[0].get('high', 0.0) if not prev_rows.empty else None

            triggered_limit_up = False
            triggered_pull_up = False
            # 漲停進場條件
            if high == limit_up_price:
                if previous_high is None or previous_high < limit_up_price:
                    triggered_limit_up = True
                    print(f"{stock_symbol} 已觸發【漲停進場】條件")
            # 拉高進場條件：除了 5min_pct_increase >=2.0 外，再加入成交量檢查
            elif five_min_increase >= 2.0:
                first_vols = df[df['time'].astype(str).isin(["09:00:00", "09:01:00", "09:02:00"])]
                avg_volume = first_vols['volume'].mean() if len(first_vols) > 0 else 0
                current_volume = row.get('volume', 0)
                if current_volume > 1.5 * avg_volume:
                    triggered_pull_up = True
                    print(f"{stock_symbol} 已觸發【拉高進場】條件 (Volume: {current_volume} > 1.5*Avg({avg_volume}))")
                else:
                    print(f"{stock_symbol} 未滿成交量條件 (Volume: {current_volume} <= 1.5*Avg({avg_volume}))")
            if triggered_limit_up or triggered_pull_up:
                eligible_stocks_for_entry.append({
                    'symbol': stock_symbol,
                    'condition': 'limit_up' if triggered_limit_up else 'pull_up',
                    'group': group
                })

    # 2. 將觸發條件的股票所屬族群標記為「觀察中」
    if eligible_stocks_for_entry:
        group_trigger = {}
        for stock in eligible_stocks_for_entry:
            grp = stock['group']
            cond = stock['condition']
            if grp not in group_trigger:
                group_trigger[grp] = set()
            group_trigger[grp].add(cond)
        for group, conditions in group_trigger.items():
            if group not in group_positions or not group_positions[group]:
                trigger_type = "漲停進場" if "limit_up" in conditions else "拉高進場"
                group_positions[group] = {
                    "status": "觀察中",
                    "start_time": datetime.combine(date.today(), current_time.time()),
                    "trigger": trigger_type,
                    "tracking_list": {}
                }
                msg = f"更新族群 {group} 狀態為觀察中，觸發方式：{trigger_type}"
                message_log.append((current_time_str, msg))
                print(msg)
    else:
        message_log.append((current_time_str, "沒有符合進場條件的股票"))

    # 2.5 若族群為【漲停進場】，持續更新追蹤清單與輸出等待狀態訊息
    now_full = datetime.combine(date.today(), current_time.time())
    for group, status in group_positions.items():
        if isinstance(status, dict) and status.get("status") == "觀察中" and status.get("trigger") == "漲停進場":
            start_time = status.get("start_time")
            elapsed = (now_full - start_time).total_seconds() / 60.0
            if "waiting_start" not in status:
                status["waiting_start"] = datetime.now()
                print(f"{group} 族群開始等待階段")
            current_tracking = status.get("tracking_list", {})
            for stock_symbol in consolidated_symbols.get(group, []):
                df = stock_data_collection.get(stock_symbol)
                if df is None or df.empty:
                    continue
                rows = df[df['time'] == trading_time]
                if rows.empty:
                    continue
                row = rows.iloc[0]
                if row.get('5min_pct_increase', 0) >= 1.5:
                    if any(item for item in eligible_stocks_for_entry if item['group'] == group and item['symbol'] == stock_symbol and item['condition'] == 'limit_up'):
                        continue
                    if stock_symbol not in current_tracking:
                        current_tracking[stock_symbol] = {'rise': row.get('rise', row.get('5min_pct_increase', 0)), 'row': row}
                        print(f"{stock_symbol} 加入到 {group} 追蹤清單")
            status["tracking_list"] = current_tracking
            if elapsed < wait_minutes:
                print(f"{group} 族群，等待第 {int(elapsed)} 分鐘")
    # 3. 對每個「觀察中」的族群（等待時間已達）進行進場判斷
    groups_to_evaluate = []
    now_full = datetime.combine(date.today(), current_time.time())
    for group, status in group_positions.items():
        if isinstance(status, dict) and status.get("status") == "觀察中":
            start_time = status.get("start_time")
            elapsed = (now_full - start_time).total_seconds() / 60.0
            if elapsed >= wait_minutes:
                groups_to_evaluate.append(group)

    idx = 0
    for group in groups_to_evaluate:
        grp_status = group_positions[group]
        trigger = grp_status.get("trigger")
        if trigger == "漲停進場":
            tracking_list = []
            for stock_symbol in consolidated_symbols.get(group, []):
                df = stock_data_collection.get(stock_symbol)
                if df is None or df.empty:
                    continue
                rows = df[df['time'] == trading_time]
                if rows.empty:
                    continue
                row = rows.iloc[0]
                if row.get('5min_pct_increase', 0) >= 1.5:
                    if any(item for item in eligible_stocks_for_entry if item['group'] == group and item['symbol'] == stock_symbol and item['condition'] == 'limit_up'):
                        continue
                    tracking_list.append({
                        'symbol': stock_symbol,
                        'rise': row.get('rise', row.get('5min_pct_increase', 0)),
                        'row': row
                    })
            grp_status["tracking_list"] = tracking_list
            if tracking_list:
                sorted_list = sorted(tracking_list, key=lambda x: x['rise'], reverse=True)
                chosen_index = len(sorted_list) // 2  # 選擇排序後中間偏後的股票
                chosen_stock = sorted_list[chosen_index]
                entry_price = chosen_stock['row'].get('close')
                if entry_price is None:
                    message_log.append((current_time_str, f"{chosen_stock['symbol']} 進場失敗，無法取得價格。"))
                    print(f"無法取得 {chosen_stock['symbol']} 的進場價格，進場失敗。")
                else:
                    # 根據 entry_trade 公式計算 shares 與其他參數
                    shares = round((capital_per_stock * 10000) / (entry_price * 1000))
                    sell_cost = shares * entry_price * 1000
                    entry_fee = int(sell_cost * (transaction_fee * 0.01) * (transaction_discount * 0.01))
                    tax = int(sell_cost * (trading_tax * 0.01))
                    # 設定區間價差與 tick_unit
                    if entry_price < 10:
                        current_price_gap = price_gap_below_50
                        tick_unit = 0.01
                    elif entry_price < 50:
                        current_price_gap = price_gap_50_to_100
                        tick_unit = 0.05
                    elif entry_price < 100:
                        current_price_gap = price_gap_50_to_100
                        tick_unit = 0.1
                    elif entry_price < 500:
                        current_price_gap = price_gap_100_to_500
                        tick_unit = 0.5
                    elif entry_price < 1000:
                        current_price_gap = price_gap_500_to_1000
                        tick_unit = 1
                    else:
                        current_price_gap = price_gap_above_1000
                        tick_unit = 5
                    # 從 chosen_stock 的資料中取得最高價
                    highest_on_entry = chosen_stock['row'].get('highest')
                    if highest_on_entry is None or highest_on_entry == 0:
                        highest_on_entry = entry_price
                    # 計算價格差額（乘以1000換算）
                    price_diff = (highest_on_entry - entry_price) * 1000
                    if price_diff < current_price_gap:
                        stop_loss_type = 'price_difference'
                        stop_loss_threshold = entry_price + (current_price_gap / 1000)
                    else:
                        stop_loss_type = 'over_high'
                        stop_loss_threshold = highest_on_entry + tick_unit
                    # 更新 current_position 變數
                    current_position = {
                        'symbol': chosen_stock['symbol'],
                        'shares': shares,
                        'entry_price': entry_price,
                        'sell_cost': sell_cost,
                        'entry_fee': entry_fee,
                        'tax': tax,
                        'entry_time': current_time_str,
                        'current_price_gap': current_price_gap,
                        'tick_unit': tick_unit,
                        'highest_on_entry': highest_on_entry,
                        'stop_loss_type': stop_loss_type,
                        'stop_loss_threshold': stop_loss_threshold
                    }
                    # 廣播進場訊息（包含停損價）
                    msg = f"{GREEN}進場！股票代號：{chosen_stock['symbol']}，進場 {shares} 張。停損價：{stop_loss_threshold:.2f}{RESET}"
                    message_log.append((current_time_str, msg))
                    print(msg)
                    in_position = True
                    group_positions[group] = "已進場"

                    #將chosen_stock['symbol']轉換為int純數字類型
                    stock_code_int = int(chosen_stock['symbol'])

                    contract = getattr(api.Contracts.Stocks.TSE, "TSE" + stock_code_int)

                    # 證券委託單 - 先賣後買市價單
                    order = api.Order(
                        price=0,                                        # 價格
                        quantity=shares,                                # 數量
                        action=sj.constant.Action.Sell,                 # 賣
                        price_type=sj.constant.StockPriceType.MKT,      # 市價單
                        order_type=sj.constant.OrderType.IOC,           # 委託條件
                        order_lot=sj.constant.StockOrderLot.Common,     # 現股
                        daytrade_short=True,                            # 先賣後買
                        account=api.stock_account                       # 下單帳號
                    )

                    # 下單
                    trade = api.place_order(contract, order)

                    #設定觸價單
                    t_cmd = tp.TouchCmd(code=f"{stock_code_int}", close=tp.Price(price=stop_loss_threshold, trend="Equal"))

                    #設定停損動作
                    o_cmd = tp.OrderCmd(
                        code=f"{stock_code_int}",
                        order=sj.Order(
                            price=0,
                            quantity=shares,
                            action="Buy",
                            order_type="ROD",
                            price_type="MKT"
                        )
                    )
                    #完成設定觸價停損
                    tcond = tp.TouchOrderCond(t_cmd, o_cmd)
                    to.add_condition(tcond)

            else:
                msg = f"漲停進場 {group} 族群無符合進場條件的股票，取消進場。"
                message_log.append((current_time_str, msg))
                print(msg)
                group_positions[group] = False

        elif trigger == "拉高進場":
            # 拉高進場邏輯：新增成交量檢查、領漲追蹤、領漲替換及等待計數修正，
            # 並分為等待前與等待後兩個階段。
            tracking_list = []
            for stock_symbol in consolidated_symbols.get(group, []):
                df = stock_data_collection.get(stock_symbol)
                if df is None or df.empty:
                    continue
                rows = df[df['time'] == trading_time]
                if rows.empty:
                    continue
                row = rows.iloc[0]
                if row.get('5min_pct_increase', 0) >= 1.5:
                    tracking_list.append({
                        'symbol': stock_symbol,
                        'rise': row.get('rise', row.get('5min_pct_increase', 0)),
                        'row': row
                    })
            grp_status["tracking_list"] = tracking_list
            if tracking_list:
                if "leader_announced" not in grp_status:
                    leader_candidate = max(tracking_list, key=lambda x: x['rise'])
                    leader_symbol = leader_candidate['symbol']
                    leader_broadcast = f"拉高進場 {group} 族群，領漲：{leader_symbol}"
                    message_log.append((current_time_str, leader_broadcast))
                    print(leader_broadcast)
                    grp_status["leader_announced"] = leader_symbol
                else:
                    leader_symbol = grp_status["leader_announced"]
                    leader_candidate = next((item for item in tracking_list if item['symbol'] == leader_symbol), None)
                    if leader_candidate is None:
                        leader_candidate = max(tracking_list, key=lambda x: x['rise'])
                        leader_symbol = leader_candidate['symbol']
                        leader_broadcast = f"拉高進場 {group} 族群，重新領漲：{leader_symbol}"
                        message_log.append((current_time_str, leader_broadcast))
                        print(leader_broadcast)
                        grp_status["leader_announced"] = leader_symbol

                leader_df = stock_data_collection.get(leader_symbol)
                if leader_df is not None and not leader_df.empty:
                    cur_leader_rows = leader_df[leader_df['time'] == trading_time]
                    prev_leader_rows = leader_df[leader_df['time'] == (datetime.combine(date.today(), trading_time) - timedelta(minutes=1)).time()]
                    if not cur_leader_rows.empty and not prev_leader_rows.empty:
                        current_high_leader = cur_leader_rows.iloc[0].get('high', 0)
                        previous_high_leader = prev_leader_rows.iloc[0].get('high', 0)
                        current_leader_rise = leader_candidate['rise']
                        if "waiting_start" not in grp_status:
                            replacement_candidate = None
                            for item in tracking_list:
                                if item['symbol'] != leader_symbol and item['rise'] > current_leader_rise:
                                    replacement_candidate = item
                                    break
                            if replacement_candidate:
                                new_leader_symbol = replacement_candidate['symbol']
                                replacement_broadcast = f"拉高進場 {group} 族群，領漲替換，新領漲：{new_leader_symbol}"
                                message_log.append((current_time_str, replacement_broadcast))
                                print(replacement_broadcast)
                                grp_status["leader_announced"] = new_leader_symbol
                                grp_status.pop("waiting_start", None)
                                grp_status.pop("waiting_counter", None)
                                grp_status.pop("leader_reversal_rise", None)
                                continue
                        if current_high_leader <= previous_high_leader:
                            if "waiting_start" not in grp_status:
                                grp_status["waiting_start"] = datetime.now()
                                grp_status["waiting_counter"] = 1
                                grp_status["leader_reversal_rise"] = current_leader_rise
                                reversal_broadcast = f"拉高進場 {group} 族群，領漲 {leader_symbol} 反轉，開始等待。"
                                message_log.append((current_time_str, reversal_broadcast))
                                print(reversal_broadcast)
                            else:
                                waiting_elapsed = int((datetime.now() - grp_status["waiting_start"]).total_seconds() / 60) + 1
                                grp_status["waiting_counter"] = waiting_elapsed
                                wait_broadcast = f"拉高進場 {group} 族群，等待第 {waiting_elapsed} 分鐘"
                                print(wait_broadcast)
                            replacement_candidate = None
                            for item in tracking_list:
                                if item['symbol'] != leader_symbol and item['rise'] > grp_status.get("leader_reversal_rise", 0):
                                    replacement_candidate = item
                                    break
                            if replacement_candidate:
                                new_leader_symbol = replacement_candidate['symbol']
                                replacement_broadcast = f"拉高進場 {group} 族群，領漲替換，新領漲：{new_leader_symbol}"
                                message_log.append((current_time_str, replacement_broadcast))
                                print(replacement_broadcast)
                                grp_status["leader_announced"] = new_leader_symbol
                                grp_status["leader_reversal_rise"] = replacement_candidate['rise']
                                grp_status["waiting_start"] = datetime.now()
                                grp_status["waiting_counter"] = 1
                                continue
                            if current_leader_rise > grp_status.get("leader_reversal_rise", 0):
                                print(f"拉高進場 {group} 族群，原領漲 {leader_symbol} 回升，取消等待，恢復正常追蹤。")
                                grp_status.pop("waiting_start", None)
                                grp_status.pop("waiting_counter", None)
                                grp_status.pop("leader_reversal_rise", None)
                            if grp_status.get("waiting_counter", 0) >= wait_minutes:
                                eligible_for_entry = []
                                for item in tracking_list:
                                    if item['symbol'] == leader_symbol:
                                        continue
                                    r = item['row']
                                    rise_val = r.get('rise', 0)
                                    current_price = r.get('close')
                                    if current_price is None or rise_val < -3 or rise_val > 7:
                                        continue
                                    eligible_for_entry.append({'symbol': item['symbol'], 'rise': rise_val, 'row': r})
                                if eligible_for_entry:
                                    eligible_sorted = sorted(eligible_for_entry, key=lambda x: x['rise'], reverse=True)
                                    chosen_index = len(eligible_sorted) // 2
                                    chosen_stock = eligible_sorted[chosen_index]
                                    entry_price = chosen_stock['row'].get('close')
                                    if entry_price is None:
                                        message_log.append((current_time_str, f"{chosen_stock['symbol']} 進場失敗，無法取得價格。"))
                                        print(f"無法取得 {chosen_stock['symbol']} 的進場價格，進場失敗。")
                                    else:
                                        shares = round((capital_per_stock * 10000) / (entry_price * 1000))
                                        sell_cost = shares * entry_price * 1000
                                        entry_fee = int(sell_cost * (transaction_fee * 0.01) * (transaction_discount * 0.01))
                                        tax = int(sell_cost * (trading_tax * 0.01))
                                        # 設定區間價差與 tick_unit
                                        if entry_price < 10:
                                            current_price_gap = price_gap_below_50
                                            tick_unit = 0.01
                                        elif entry_price < 50:
                                            current_price_gap = price_gap_50_to_100
                                            tick_unit = 0.05
                                        elif entry_price < 100:
                                            current_price_gap = price_gap_50_to_100
                                            tick_unit = 0.1
                                        elif entry_price < 500:
                                            current_price_gap = price_gap_100_to_500
                                            tick_unit = 0.5
                                        elif entry_price < 1000:
                                            current_price_gap = price_gap_500_to_1000
                                            tick_unit = 1
                                        else:
                                            current_price_gap = price_gap_above_1000
                                            tick_unit = 5
                                        highest_on_entry = chosen_stock['row'].get('highest')
                                        if highest_on_entry is None or highest_on_entry == 0:
                                            highest_on_entry = entry_price
                                        price_diff = (highest_on_entry - entry_price) * 1000
                                        if price_diff < current_price_gap:
                                            stop_loss_type = 'price_difference'
                                            stop_loss_threshold = entry_price + (current_price_gap / 1000)
                                        else:
                                            stop_loss_type = 'over_high'
                                            stop_loss_threshold = highest_on_entry + tick_unit
                                        current_position = {
                                            'symbol': chosen_stock['symbol'],
                                            'shares': shares,
                                            'entry_price': entry_price,
                                            'sell_cost': sell_cost,
                                            'entry_fee': entry_fee,
                                            'tax': tax,
                                            'entry_time': current_time_str,
                                            'current_price_gap': current_price_gap,
                                            'tick_unit': tick_unit,
                                            'highest_on_entry': highest_on_entry,
                                            'stop_loss_type': stop_loss_type,
                                            'stop_loss_threshold': stop_loss_threshold
                                        }
                                        entry_broadcast = f"{GREEN}進場！股票代號：{chosen_stock['symbol']}，進場 {shares} 張。停損價：{stop_loss_threshold:.2f}{RESET}"
                                        message_log.append((current_time_str, entry_broadcast))
                                        print(entry_broadcast)
                                        in_position = True
                                        group_positions[group] = "已進場"

                                        stock_code_int = int(chosen_stock['symbol'])
                                        contract = getattr(api.Contracts.Stocks.TSE, "TSE" + stock_code_int)
                    
                                        # 證券委託單 - 先賣後買市價單
                                        order = api.Order(
                                            price=0,                                        # 價格
                                            quantity=shares,                                # 數量
                                            action=sj.constant.Action.Sell,                 # 賣
                                            price_type=sj.constant.StockPriceType.MKT,      # 市價單
                                            order_type=sj.constant.OrderType.IOC,           # 委託條件
                                            order_lot=sj.constant.StockOrderLot.Common,     # 現股
                                            daytrade_short=True,                            # 先賣後買
                                            account=api.stock_account                       # 下單帳號
                                        )

                                        trade = api.place_order(contract,order)
                                        #設定觸價單
                                        t_cmd = tp.TouchCmd(
                                            code=f"{stock_code_int}",
                                            close=tp.Price(price=stop_loss_threshold, trend="Equal")
                                        )
                                        #設定停損動作
                                        o_cmd = tp.OrderCmd(
                                            code=f"{stock_code_int}",
                                            order=sj.Order(
                                                price=0,
                                                quantity=shares,
                                                action="Buy",
                                                order_type="ROD",
                                                price_type="MKT"
                                            )
                                        )
                                        #完成設定觸價停損
                                        tcond = tp.TouchOrderCond(t_cmd, o_cmd)
                                        to.add_condition(tcond)
                                else:
                                    cancel_broadcast = f"拉高進場 {group} 族群無符合進場條件的股票，取消進場。"
                                    message_log.append((current_time_str, cancel_broadcast))
                                    print(cancel_broadcast)
                                    group_positions[group] = False
                        else:
                            replacement_candidate = None
                            for item in tracking_list:
                                if item['symbol'] != leader_symbol and item['rise'] > current_leader_rise:
                                    replacement_candidate = item
                                    break
                            if replacement_candidate:
                                new_leader_symbol = replacement_candidate['symbol']
                                replacement_broadcast = f"拉高進場 {group} 族群，領漲替換，新領漲：{new_leader_symbol}"
                                message_log.append((current_time_str, replacement_broadcast))
                                print(replacement_broadcast)
                                grp_status["leader_announced"] = new_leader_symbol
                                grp_status.pop("waiting_start", None)
                                grp_status.pop("waiting_counter", None)
                                grp_status.pop("leader_reversal_rise", None)
                    else:
                        print(f"無法取得 {leader_symbol} 的完整K線數據，跳過反轉判斷。")
                else:
                    print(f"{group} 族群的追蹤清單為空，無法評估進場。")
        idx += 1

    message_log.sort(key=lambda x: str(x[0]))
    for log_time, msg in message_log:
        print(f"[{log_time}] {msg}")
    message_log.clear()

#盤中13:30出場
def exit_trade_live():
    """
    此函數依據設定，於 13:26 時進行出場動作：
      1. 從全域變數 to 中取得所有尚存的觸價委託單（to.conditions）
         -> 條件字典的 key 為股票代號（例如 "2330"），value 為該股票所有委託單列表
      2. 依據每個股票代號的所有委託單，累加取出進場張數（quantity），形成 exit_data 字典，
         格式如 { "2330": 10, "2317": 12, ... }
      3. 將 exit_data 寫入本地檔案 "enter_exit.json"
      4. 重新讀取 "enter_exit.json" 的資料
      5. 對 exit_data 中每一筆資料，利用股票代號與進場張數建立出場委託單：
         - 利用 getattr 與字串串接組成 "TSE" + stock_code，取得該股票的 contract
         - 以 contract.limit_up 作為價格，下單一筆 ROC 條件的限價單 (LMT) 買進委託
         - 呼叫 api.place_order 發出出場委託單，並打印下單資訊
      6. 刪除所有尚存的觸價委託單
    """
    # 取得所有尚存的觸價委託單 (to.conditions)
    conditions_dict = to.conditions
    exit_data = {}
    # 遍歷每個股票代號及其委託單列表，累加進場張數
    for stock_code, cond_list in conditions_dict.items():
        total_quantity = 0
        for cond in cond_list:
            # 假設每個觸價委託單物件 cond 其 order 屬性中有 quantity 屬性
            try:
                qty = getattr(cond.order, 'quantity', 0)
                total_quantity += int(qty)
            except Exception as e:
                print(f"讀取股票 {stock_code} 的數量時發生錯誤：{e}")
        if total_quantity > 0:
            exit_data[stock_code] = total_quantity

    # 將 exit_data 寫入 "enter_exit.json"
    try:
        with open("enter_exit.json", "w", encoding="utf-8") as f:
            json.dump(exit_data, f, ensure_ascii=False, indent=4)
        print("已將當前觸價委託單的股票代號和進場張數儲存至 enter_exit.json:")
        print(exit_data)
    except Exception as e:
        print(f"寫入 enter_exit.json 檔案失敗：{e}")
        return

    # 讀取最新的 exit data
    try:
        with open("enter_exit.json", "r", encoding="utf-8") as f:
            exit_info = json.load(f)
    except Exception as e:
        print(f"讀取 enter_exit.json 檔案失敗：{e}")
        return

    if not exit_info:
        print("enter_exit.json 中沒有觸價委託單資料，終止出場程序。")
        return

    # 對每筆 exit_info 中的資料，建立出場委託單
    for stock_code, shares in exit_info.items():
        try:
            # 取得 contract 物件，例如將 "TSE" 與 stock_code 串接，得到 "TSE2330"
            contract = getattr(api.Contracts.Stocks.TSE, "TSE" + stock_code)
            # 取得漲停價
            limit_up_price = contract.limit_up

            # 建立限價買進的委託單 (ROC 條件)
            order = api.Order(
                action=sj.constant.Action.Buy,
                price=limit_up_price,
                quantity=shares,
                price_type=sj.constant.StockPriceType.LMT,
                order_type=sj.constant.OrderType.ROC,
                order_lot=sj.constant.StockOrderLot.Common,
                account=api.stock_account
            )
            # 下單出場委託單
            trade = api.place_order(contract, order)
            print(f"下單出場：股票 {stock_code}，數量 {shares} 張；價格設定為漲停價 {limit_up_price}")
        except Exception as e:
            print(f"處理股票 {stock_code} 時發生錯誤：{e}")

    # 刪除所有尚存的觸價委託單
    for stock_code, cond_list in list(conditions_dict.items()):
        for cond in cond_list:
            try:
                to.delete_condition(cond)
            except Exception as e:
                print(f"刪除股票 {stock_code} 的觸價委託單時發生錯誤：{e}")

    print("出場委託單已全部下單，並刪除所有觸價委託單。")

def wait_until_next_minute():
    now = datetime.now()
    next_minute = (now + timedelta(minutes=1)).replace(second=5, microsecond=0)
    wait_seconds = (next_minute - now).total_seconds()
    time_module.sleep(wait_seconds)

def user_wants_to_quit():
    if msvcrt.kbhit():
        key = msvcrt.getwch()
        if key.upper() == 'Q':
            while msvcrt.kbhit():
                msvcrt.getwch()
            return True
    return False
    
def truncate_to_two_decimals(value):
    if isinstance(value, float):
        return math.floor(value * 100) / 100
    return value

def calculate_5min_pct_increase(new_candle, existing_candles):
    new_candle['5min_pct_increase'] = 0.0
    all_candles = existing_candles + [new_candle]
    num_existing_candles = len(existing_candles)
    if num_existing_candles == 0:
        new_candle['5min_pct_increase'] = 0.0
    else:
        if num_existing_candles < 4:
            relevant_candles = all_candles
        else:
            relevant_candles = existing_candles[-4:] + [new_candle]

        close_prices = [float(c['close']) for c in relevant_candles if c.get('close') is not None]

        if len(close_prices) < 2:
            new_candle['5min_pct_increase'] = 0.0
        else:
            max_close = max(close_prices)
            min_close = min(close_prices)
            index_max = close_prices.index(max_close)
            index_min = close_prices.index(min_close)

            if index_max > index_min:
                pct_increase = ((max_close - min_close) / min_close) * 100
            else:
                pct_increase = ((min_close - max_close) / max_close) * 100

            new_candle['5min_pct_increase'] = round(pct_increase, 2)
    return new_candle

def save_auto_intraday_data(auto_intraday_data):
    try:
        with open('auto_intraday.json', 'w', encoding='utf-8') as f:
            json.dump(auto_intraday_data, f, ensure_ascii=False, indent=4, default=str)
        print("已成功儲存 auto_intraday.json")
    except Exception as e:
        print(f"儲存 auto_intraday.json 時發生錯誤：{e}")

def update_kline_data_menu():
    while True:
        print("\n更新K線數據選單：")
        print("1. 更新K線數據")
        print("2. 查看K線數據")
        print("0. 返回主選單")
        choice = input("請輸入選項：")
        if choice == '1':
            update_kline_data()
        elif choice == '2':
            view_kline_data()
        elif choice == '0':
            main_menu()
        else:
            print("無效的選項，請重新輸入")

def convert_datetime_to_str(obj):
    if isinstance(obj, dict):
        return {k: convert_datetime_to_str(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_datetime_to_str(element) for element in obj]
    elif isinstance(obj, (datetime, pd.Timestamp, time)):
        return obj.isoformat()
    else:
        return obj

def update_kline_data():
    client, api_key = init_fugle_client()
    matrix_dict_analysis = load_matrix_dict_analysis()
    if not matrix_dict_analysis:
        print("沒有任何族群資料，請先管理族群。")
        return

    print("正在更新處置股清單...")
    fetch_disposition_stocks(client, matrix_dict_analysis)
    print("處置股清單已更新。")

    disposition_stocks = load_disposition_stocks()
    intraday_kline_data = {}

    count = 0
    for group, symbols in matrix_dict_analysis.items():
        print(f"處理族群: {group}")
        filtered_symbols = [symbol for symbol in symbols if symbol not in disposition_stocks]

        if not filtered_symbols:
            print(f"族群 {group} 過濾後沒有任何可供分析的股票。")
            continue

        for symbol in filtered_symbols:
            if count >= 55:
                print("已達到55次API請求，休息1分鐘...")
                time_module.sleep(60)
                count = 0

            daily_kline_df = fetch_daily_kline_data(client, symbol, days=2)
            count += 1

            if daily_kline_df.empty:
                print(f"無法取得 {symbol} 的日K數據，跳過。")
                continue

            sorted_daily_data = sorted(daily_kline_df.to_dict(orient='records'), key=lambda x: x['date'], reverse=True)
            if len(sorted_daily_data) > 1:
                yesterday_close_price = sorted_daily_data[1].get('close', 0)
            else:
                print(f"警告：{symbol} 的日K數據不足以提取上一個交易日的收盤價。")
                yesterday_close_price = 0

            if count >= 55:
                print("已達到55次API請求，休息1分鐘...")
                time_module.sleep(60)
                count = 0

            recent_day = get_recent_trading_day()
            current_time = datetime.now()
            today = current_time.date()
            if isinstance(recent_day, str):
                try:
                    trading_day_date = datetime.strptime(recent_day, '%Y-%m-%d').date()
                except ValueError as ve:
                    print(f"日期格式錯誤：{recent_day}，錯誤訊息：{ve}")
                    continue
            elif isinstance(recent_day, datetime):
                trading_day_date = recent_day.date()
            elif isinstance(recent_day, date):
                trading_day_date = recent_day
            else:
                print(f"未知的 recent_day 類型：{type(recent_day)}，值：{recent_day}")
                continue

            if trading_day_date < today:
                initial_fetch_end_time_str = "13:30"
            else:
                market_end_time = current_time.replace(hour=13, minute=30, second=0, microsecond=0)
                if current_time > market_end_time:
                    initial_fetch_end_time_str = "13:30"
                else:
                    initial_fetch_end_time = (current_time - timedelta(minutes=1)).replace(second=0, microsecond=0)
                    initial_fetch_end_time_str = initial_fetch_end_time.strftime('%H:%M')

            print(f"正在取得 {symbol} 的一分K數據從 09:00 到 {initial_fetch_end_time_str}...")

            intraday_df = fetch_intraday_data(
                client=client,
                symbol=symbol,
                trading_day=recent_day,
                yesterday_close_price=yesterday_close_price,
                start_time="09:00",
                end_time=initial_fetch_end_time_str
            )
            count += 1

            if intraday_df.empty:
                print(f"無法取得 {symbol} 的一分K數據，跳過。")
                continue
            intraday_df = calculate_5min_pct_increase_and_highest(intraday_df)

            intraday_data = intraday_df.to_dict(orient='records')

            intraday_kline_data[symbol] = intraday_data
            print(f"已取得 {symbol} 的一分K數據並加入 intraday_kline_data.json")

    intraday_kline_data_str = convert_datetime_to_str(intraday_kline_data)
    with open('intraday_kline_data.json', 'w', encoding='utf-8') as f:
        json.dump(intraday_kline_data_str, f, indent=4, ensure_ascii=False, default=str)
    print("K線數據已儲存。")

    mt_matrix_dict = {}
    for group, symbols in matrix_dict_analysis.items():
        stock_data_list = []
        for symbol in symbols:
            if symbol in intraday_kline_data:
                df = pd.DataFrame(intraday_kline_data[symbol])
                if 'symbol' not in df.columns:
                    df['symbol'] = symbol
                stock_data_list.append(df)

        if stock_data_list:
            print(f"正在計算族群 {group} 的相似度...")
            similarity_df = calculate_kline_similarity(stock_data_list)
            similarity_df = similarity_df[similarity_df['similarity_score'] > 0.3]

            if similarity_df.empty:
                print(f"族群 {group} 沒有相似度大於0.3 的股票組合。")
                continue
            similarity_records = similarity_df.to_dict(orient='records')

            for record in similarity_records:
                record['group'] = group

            mt_matrix_dict[group] = similarity_records
            print(f"族群 {group} 的相似度計算完成並加入 mt_matrix_dict。")

    with open('mt_matrix_dict.json', 'w', encoding='utf-8') as f:
        json.dump(mt_matrix_dict, f, indent=4, ensure_ascii=False, default=str)

    print("相似度計算完成並已儲存至 mt_matrix_dict.json。")

    consolidate_and_save_stock_symbols()
    print("股票代號已統整並儲存至 nb_matrix_dict.json，按族群分類。")

    print("K線數據更新完成。")

def view_kline_data():
    if not os.path.exists('intraday_kline_data.json'):
        print("尚未更新一分K數據，請先更新K線數據。")
        return
    with open('intraday_kline_data.json', 'r', encoding='utf-8') as f:
        intraday_kline_data = json.load(f)
    
    for symbol, data in intraday_kline_data.items():
        print(f"\n股票代號：{symbol} 的一分K數據：")
        df = pd.DataFrame(data)
        if df.empty:
            print("沒有資料。")
            continue
        
        if 'time' in df.columns:
            try:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore", UserWarning)
                    df['time'] = pd.to_datetime(df['time'])
            except Exception as e:
                print(f"轉換時間欄位時發生錯誤：{e}")
                continue
        
        print(df)

def save_settings():
    with open('settings.json', 'w', encoding='utf-8') as f:
        json.dump({
            'capital_per_stock': capital_per_stock,
            'transaction_fee': transaction_fee,
            'transaction_discount': transaction_discount,
            'trading_tax': trading_tax,
            'below_50': below_50,
            'price_gap_50_to_100': price_gap_50_to_100,
            'price_gap_100_to_500': price_gap_100_to_500,
            'price_gap_500_to_1000': price_gap_500_to_1000,
            'price_gap_above_1000': price_gap_above_1000,
            'allow_reentry_after_stop_loss': allow_reentry_after_stop_loss
        }, f, indent=4)

def load_settings():
    global capital_per_stock, transaction_fee, transaction_discount, trading_tax
    global below_50, price_gap_50_to_100, price_gap_100_to_500, price_gap_500_to_1000, price_gap_above_1000
    global allow_reentry_after_stop_loss
    if os.path.exists('settings.json'):
        with open('settings.json', 'r', encoding='utf-8') as f:
            settings = json.load(f)
            capital_per_stock = settings.get('capital_per_stock', 0)
            transaction_fee = settings.get('transaction_fee', 0)
            transaction_discount = settings.get('transaction_discount', 0)
            trading_tax = settings.get('trading_tax', 0)
            below_50 = settings.get('below_50', 0)
            price_gap_50_to_100 = settings.get('price_gap_50_to_100', 0)
            price_gap_100_to_500 = settings.get('price_gap_100_to_500', 0)
            price_gap_500_to_1000 = settings.get('price_gap_500_to_1000', 0)
            price_gap_above_1000 = settings.get('price_gap_above_1000', 0)
            allow_reentry_after_stop_loss = settings.get('allow_reentry_after_stop_loss', False)
    else:
        capital_per_stock = 1000
        transaction_fee = 0.1425
        transaction_discount = 20.0
        trading_tax = 0.15
        below_50 = 500
        price_gap_50_to_100 = 1000
        price_gap_100_to_500 = 2000
        price_gap_500_to_1000 = 3000
        price_gap_above_1000 = 5000
        allow_reentry_after_stop_loss = False

def settings_menu():
    global capital_per_stock, transaction_fee, transaction_discount, trading_tax
    global below_50, price_gap_50_to_100, price_gap_100_to_500, price_gap_500_to_1000, price_gap_above_1000
    global allow_reentry_after_stop_loss
    while True:
        print("\n設定選單：")
        print(f"1. 設定每檔股票投入資本額（目前為 {capital_per_stock} 萬元）")
        print(f"2. 手續費設定，目前為 {transaction_fee}%")
        print(f"3. 手續費折數設定，目前為 {transaction_discount}%")
        print(f"4. 證交稅設定，目前為 {trading_tax}%")
        print("5. 價差停損設定")
        print("6. 停損再進場設定")
        print("0. 返回主選單")
        choice = input("請輸入選項：")
        if choice == "1":
            set_capital_per_stock()
        elif choice == "2":
            transaction_fee = float(input("請輸入手續費（%）："))
            save_settings()
        elif choice == "3":
            transaction_discount = float(input("請輸入手續費折數（%）："))
            save_settings()
        elif choice == "4":
            trading_tax = float(input("請輸入證交稅（%）："))
            save_settings()
        elif choice == "5":
            price_gap_stop_loss_menu()
        elif choice == "6":
            stop_loss_reentry_menu()
        elif choice == "0":
            main_menu()
        else:
            print("無效的選項，請重新輸入")

def stop_loss_reentry_menu():
    global allow_reentry_after_stop_loss
    while True:
        status = "開啟" if allow_reentry_after_stop_loss else "關閉"
        print(f"\n目前為({status}停損後進場)")
        print("1.開啟停損後進場")
        print("2.關閉停損後進場")
        print("3.返回上一頁")
        choice = input("請輸入選項：")
        if choice == '1':
            allow_reentry_after_stop_loss = True
            print("已開啟停損後進場功能")
            save_settings()
        elif choice == '2':
            allow_reentry_after_stop_loss = False
            print("已關閉停損後進場功能")
            save_settings()
        elif choice == '3':
            settings_menu()
        else:
            print("無效的選項，請重新輸入")

def price_gap_stop_loss_menu():
    global below_50, price_gap_50_to_100, price_gap_100_to_500, price_gap_500_to_1000, price_gap_above_1000
    while True:
        print(f"1. 50元以下股票停損價差，目前為 {below_50} 元")
        print(f"2. 50~100元股票停損價差，目前為 {price_gap_50_to_100} 元")
        print(f"3. 100~500元股票停損價差，目前為 {price_gap_100_to_500} 元")
        print(f"4. 500~1000元股票停損價差，目前為 {price_gap_500_to_1000} 元")
        print(f"5. 1000元以上股票停損價差，目前為 {price_gap_above_1000} 元")
        print("6. 返回上一頁")
        choice = input("請選擇要設定的項目：")
        if choice == "1":
            below_50 = float(input("請輸入50元以下股票的停損價差："))
        elif choice == "2":
            price_gap_50_to_100 = float(input("請輸入50~100元股票的停損價差："))
        elif choice == "3":
            price_gap_100_to_500 = float(input("請輸入100~500元股票的停損價差："))
        elif choice == "4":
            price_gap_500_to_1000 = float(input("請輸入500~1000元股票的停損價差："))
        elif choice == "5":
            price_gap_above_1000 = float(input("請輸入1000元以上股票的停損價差："))
        elif choice == "6":
            break
        else:
            print("無效選擇，請重試。")
        save_settings()

def simulate_trading_menu():
    matrix_dict_analysis = load_matrix_dict_analysis()
    if not matrix_dict_analysis:
        print("沒有族群資料，請先管理族群。")
        return

    while True:
        print("請選擇操作：")
        print("1. 分析單一族群")
        print("2. 分析全部族群")
        print("0. 返回主選單")
        choice = input("請輸入選項編號：")

        if choice == '1':
            group_name = input("請輸入要分析的族群名稱：")
            if group_name not in matrix_dict_analysis:
                print("沒有此族群資料")
                continue

            try:
                wait_minutes = int(input("請輸入等待時間（分鐘）："))
            except ValueError:
                print("等待時間必須是整數。")
                continue

            hold_minutes_input = input("請輸入持有時間（分鐘，輸入 'F' 代表持有到13:30強制出場）：")
            if hold_minutes_input.upper() == 'F':
                hold_minutes = None
            else:
                try:
                    hold_minutes = int(hold_minutes_input)
                except ValueError:
                    print("持有時間必須是整數或 'F'。")
                    continue

            disposition_stocks = load_disposition_stocks()
            symbols_to_analyze = matrix_dict_analysis[group_name]
            symbols_to_analyze = [symbol for symbol in symbols_to_analyze if symbol not in disposition_stocks]
            if len(symbols_to_analyze) == 0:
                print(f"{group_name} 中沒有可供分析的股票。")
                continue

            daily_kline_data, intraday_kline_data = load_kline_data()

            stock_data_collection = initialize_stock_data(symbols_to_analyze, daily_kline_data, intraday_kline_data)
            if not stock_data_collection:
                print("無法獲取有效的一分 K 資料，無法進行分析")
                continue

            total_profit, avg_profit_rate = process_group_data(stock_data_collection, wait_minutes, hold_minutes, matrix_dict_analysis, verbose=True)

            print(f"\n模擬交易完成，總利潤：{int(total_profit) if total_profit is not None else 0} 元，平均報酬率：{avg_profit_rate if avg_profit_rate is not None else 0:.2f}%\n")

        elif choice == '2':
            try:
                wait_minutes = int(input("請輸入等待時間（分鐘）："))
            except ValueError:
                print("等待時間必須是整數。")
                continue

            hold_minutes_input = input("請輸入持有時間（分鐘，輸入 'F' 代表持有到13:30強制出場）：")
            if hold_minutes_input.upper() == 'F':
                hold_minutes = None
            else:
                try:
                    hold_minutes = int(hold_minutes_input)
                except ValueError:
                    print("持有時間必須是整數或 'F'。")
                    continue

            day_total_profit = 0
            day_avg_profit_rates = []

            for group_name in matrix_dict_analysis.keys():
                print(f"\n正在分析族群：{group_name}")

                disposition_stocks = load_disposition_stocks()
                symbols_to_analyze = matrix_dict_analysis[group_name]
                symbols_to_analyze = [symbol for symbol in symbols_to_analyze if symbol not in disposition_stocks]
                if len(symbols_to_analyze) == 0:
                    print(f"{group_name} 中沒有可供分析的股票。")
                    continue

                daily_kline_data, intraday_kline_data = load_kline_data()

                stock_data_collection = initialize_stock_data(symbols_to_analyze, daily_kline_data, intraday_kline_data)
                if not stock_data_collection:
                    print(f"無法獲取 {group_name} 的有效一分 K 資料，跳過。")
                    continue

                total_profit, avg_profit_rate = process_group_data(stock_data_collection, wait_minutes, hold_minutes, matrix_dict_analysis, verbose=True)

                if total_profit is not None and avg_profit_rate is not None:
                    day_total_profit += total_profit
                    day_avg_profit_rates.append(avg_profit_rate)
                else:
                    pass

            if day_avg_profit_rates:
                day_avg_profit_rate = sum(day_avg_profit_rates) / len(day_avg_profit_rates)
            else:
                day_avg_profit_rate = 0.0

            if day_total_profit > 0:
                print(f"{RED}=" * 50)
                print(f"{RED}\n當日總利潤：{int(day_total_profit)} 元{RESET}")
                print(f"{RED}當日報酬率：{day_avg_profit_rate:.2f}%\n{RESET}")
                print(f"{RED}=" * 50)
            elif day_total_profit < 0:
                print(f"{GREEN}=" * 50)
                print(f"{GREEN}\n當日總利潤：{int(day_total_profit)} 元{RESET}")
                print(f"{GREEN}當日報酬率：{day_avg_profit_rate:.2f}%\n{RESET}")
                print(f"{GREEN}=" * 50)
            else:
                print("=" * 50)
                print(f"\n當日總利潤：{int(day_total_profit)} 元")
                print(f"當日報酬率：{day_avg_profit_rate:.2f}%\n")
                print("=" * 50)

        elif choice == '0':
            break
        else:
            print("無效的選項，請重新輸入。")

def display_disposition_stocks():
    disposition_file = 'Disposition.json'
    try:
        with open(disposition_file, 'r', encoding='utf-8') as f:
            disposition_data = json.load(f)
            if isinstance(disposition_data, list):
                stock_codes = disposition_data
            elif isinstance(disposition_data, dict):
                stock_codes = disposition_data.get("stock_codes", [])
            else:
                print(f"錯誤：{disposition_file} 文件格式不正確。")
                return
    except FileNotFoundError:
        print(f"錯誤：無法找到 {disposition_file} 文件。")
        return
    except json.JSONDecodeError:
        print(f"錯誤：{disposition_file} 文件格式不正確。")
        return

    if not stock_codes:
        print(f"{disposition_file} 中沒有任何股票代號。")
        return

    items_per_page = 10
    total_items = len(stock_codes)
    total_pages = (total_items + items_per_page - 1) // items_per_page
    current_page = 1

    while True:
        start_idx = (current_page - 1) * items_per_page
        end_idx = start_idx + items_per_page
        page_items = stock_codes[start_idx:end_idx]

        print("\n" + "=" * 50)
        print(f"{disposition_file} 股票代號列表 - 第 {current_page} 頁 / 共 {total_pages} 頁")
        print("=" * 50)
        for idx, code in enumerate(page_items, start=1 + start_idx):
            print(f"{idx}. {code}")
        print("=" * 50)
        if total_pages == 1:
            print("已顯示所有股票代號。")
            break

        print("導航選項：")
        if current_page > 1:
            print("P - 上一頁")
        if current_page < total_pages:
            print("N - 下一頁")
        print("0 - 返回主選單")

        choice = input("請輸入選項（N/P/0）：").strip().upper()

        if choice == 'N' and current_page < total_pages:
            current_page += 1
        elif choice == 'P' and current_page > 1:
            current_page -= 1
        elif choice == '0':
            break
        else:
            print("無效的選項，請重新輸入。")

def set_capital_per_stock():
    global capital_per_stock
    capital_per_stock = int(input("請輸入每檔投入資本額（萬元）："))
    print(f"每檔投入資本額已設定為：{capital_per_stock} 萬元")
    save_settings()

def maximize_profit_analysis():
    print("進入極大化利潤模式...")
    
    matrix_dict_analysis = load_matrix_dict_analysis()
    if not matrix_dict_analysis:
        print("沒有族群資料，請先管理族群。")
        return

    group_name = input("請輸入要分析的族群名稱：")
    
    if group_name not in matrix_dict_analysis:
        print("沒有此族群資料")
        return
    wait_minutes_start = int(input("請輸入等待時間起始值（分鐘）："))
    wait_minutes_end = int(input("請輸入等待時間結束值（分鐘）："))
    hold_minutes_start = int(input("請輸入持有時間起始值（分鐘，輸入0代表F）："))
    hold_minutes_end = int(input("請輸入持有時間結束值（分鐘，輸入0代表F）："))

    wait_minutes_range = range(wait_minutes_start, wait_minutes_end + 1)
    hold_minutes_range = range(hold_minutes_start, hold_minutes_end + 1)

    disposition_stocks = load_disposition_stocks()
    symbols_to_analyze = matrix_dict_analysis[group_name]
    symbols_to_analyze = [symbol for symbol in symbols_to_analyze if symbol not in disposition_stocks]
    if len(symbols_to_analyze) == 0:
        print(f"{group_name} 中沒有可供分析的股票。")
        return

    daily_kline_data, intraday_kline_data = load_kline_data()

    stock_data_collection = initialize_stock_data(symbols_to_analyze, daily_kline_data, intraday_kline_data)
    if not stock_data_collection:
        print("無法獲取有效的一分 K 資料，無法進行分析")
        return

    results_df = pd.DataFrame(columns=['等待時間', '持有時間', '總利潤', '平均報酬率'])
    results_df = results_df.astype({
        '等待時間': 'int',
        '持有時間': 'object',
        '總利潤': 'float',
        '平均報酬率': 'float'
    })

    for wait_minutes in wait_minutes_range:
        for hold_minutes in hold_minutes_range:
            hold_minutes_value = None if hold_minutes == 0 else hold_minutes
            print(f"正在分析：等待時間 {wait_minutes} 分鐘、持有時間 {'F' if hold_minutes_value is None else hold_minutes_value} 分鐘")
            
            total_profit, avg_profit_rate = process_group_data(
                stock_data_collection, wait_minutes, hold_minutes_value, matrix_dict_analysis, verbose=False)
            
            if total_profit is None:
                total_profit = 0.0
            if avg_profit_rate is None:
                avg_profit_rate = 0.0
            
            new_row = pd.DataFrame([{
                '等待時間': wait_minutes,
                '持有時間': 'F' if hold_minutes_value is None else hold_minutes_value,
                '總利潤': float(total_profit),
                '平均報酬率': float(avg_profit_rate)
            }])
            results_df = pd.concat([results_df, new_row], ignore_index=True)

    if results_df.empty:
        print("模擬結果為空，無法進行後續分析。")
        return

    max_profit = results_df['總利潤'].max()
    min_profit = results_df['總利潤'].min()
    best_combination = results_df.loc[results_df['總利潤'].idxmax()]

    print("\n利潤最大的組合：")
    print(f"等待時間：{best_combination['等待時間']} 分鐘，持有時間：{best_combination['持有時間']} 分鐘，總利潤：{int(best_combination['總利潤'])} 元，平均報酬率：{best_combination['平均報酬率']:.2f}%\n")

    pivot_df = results_df.pivot(index='等待時間', columns='持有時間', values='總利潤')

    formatted_pivot_df = pivot_df.copy()
    for col in formatted_pivot_df.columns:
        if col != '等待時間':
            formatted_pivot_df[col] = formatted_pivot_df[col].apply(lambda x: f"{int(x):,}" if pd.notnull(x) else "")

    formatted_pivot_df_reset = formatted_pivot_df.reset_index()

    print("模擬結果：")
    print(tabulate(formatted_pivot_df_reset, headers='keys', tablefmt='psql', showindex=False))

    try:
        with pd.ExcelWriter('模擬結果.xlsx', engine='openpyxl') as writer:
            pivot_df.to_excel(writer, sheet_name='模擬結果', index=True)
            workbook = writer.book
            worksheet = writer.sheets['模擬結果']
            
            max_profit = pivot_df.max().max()
            min_profit = pivot_df.min().min()

            max_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            min_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

            for row in worksheet.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    if cell.value == max_profit:
                        cell.fill = max_fill
                    elif cell.value == min_profit:
                        cell.fill = min_fill
        print("\n模擬結果已成功寫入 '模擬結果.xlsx'。")
    except Exception as e:
        print(f"\n寫入 Excel 時發生錯誤：{e}")

def manage_groups():
    current_page = 0
    page_size = 5
    groups = load_matrix_dict_analysis()
    total_pages = (len(groups) + page_size - 1) // page_size

    def display_page(page):
        start = page * page_size
        end = start + page_size
        print("=" * 50)
        print(f"族群及個股列表 - 第 {page + 1} 頁 / 共 {total_pages} 頁")
        print("=" * 50)
        for idx, (group, stocks) in enumerate(list(groups.items())[start:end], start=1):
            print(f"族群: {group}")
            for stock_idx, stock in enumerate(stocks, start=1):
                print(f"  {str(stock_idx).rjust(2)}. {stock}")
            print("-" * 50)
        print("=" * 50)
        if current_page == total_pages - 1:
            print("已顯示所有族群及個股。")
        print("=" * 50)

    while True:
        display_page(current_page)
        print("\nP：上一頁、Q：下一頁、1：新增族群/個股；、2：刪除族群/個股、0：返回主選單")
        choice = input("請選擇操作: ")

        if choice == "P":
            if current_page > 0:
                current_page -= 1
            else:
                print("已經是第一頁！")
        elif choice == "Q":
            if current_page < total_pages - 1:
                current_page += 1
            else:
                print("已經是最後一頁！")
        elif choice == "1":
            add_group_or_stock(groups)
        elif choice == "2":
            delete_group_or_stock(groups)
        elif choice == "0":
            save_matrix_dict(groups)
            break
        else:
            print("無效選項，請重新選擇。")

def add_group_or_stock(groups):
    print("\n==============================")
    print("1：新增族群、2：新增族群中的個股、3：返回選單")
    print("\n==============================")
    choice = input("請選擇操作: ").strip()

    if choice == "1":
        new_group = input("輸入新族群名稱: ").strip()
        if not new_group:
            print("族群名稱不能為空。")
            add_group_or_stock(groups)
        if new_group in groups:
            print(f"族群 '{new_group}' 已存在。")
        else:
            groups[new_group] = []
            print(f"族群 '{new_group}' 新增成功。")
    
    elif choice == "2":
        group_name = input("輸入要新增個股的族群名稱: ").strip()
        if not group_name:
            print("族群名稱不能為空。")
            add_group_or_stock(groups)
        if group_name in groups:
            current_stocks = groups[group_name]
            print(f"\n==============================")
            print(f"族群 '{group_name}' 中目前的個股:")
            if current_stocks:
                for idx, stock in enumerate(current_stocks, start=1):
                    print(f"  {str(idx).rjust(2)}. {stock}")
            else:
                print("  無")
            print("==============================\n")
            
            print(f"開始新增個股到族群 '{group_name}'。")
            print("請輸入個股代號，輸入 'Q' 以退出新增模式。")
            
            while True:
                new_stock = input("輸入個股代號 (或 'Q' 退出): ").strip()
                if new_stock.upper() == "Q":
                    print("退出新增個股模式。")
                    break
                elif not new_stock:
                    print("輸入無效，請重新輸入。")
                    continue
                elif new_stock in groups[group_name]:
                    print(f"個股 '{new_stock}' 已存在於族群 '{group_name}' 中。")
                else:
                    groups[group_name].append(new_stock)
                    print(f"個股 '{new_stock}' 已新增至族群 '{group_name}'。")
        else:
            print(f"族群 '{group_name}' 不存在。")
    
    elif choice == "0":
        print("返回主選單。")
        manage_groups()

    else:
        print("無效的選項，請重新選擇。")

def delete_group_or_stock(groups):
    print("\n==============================")
    print("1：刪除族群、2：刪除族群中的個股、3：返回選單")
    print("\n==============================")
    choice = input("請選擇操作: ").strip()

    if choice == "1":
        group_name = input("輸入要刪除的族群名稱: ").strip()
        if not group_name:
            print("族群名稱不能為空。")
            delete_group_or_stock(groups)
        if group_name in groups:
            confirm = input(f"確定要刪除族群 '{group_name}' 嗎？ (Y/N): ").strip().upper()
            if confirm == "Y":
                del groups[group_name]
                print(f"族群 '{group_name}' 已刪除。")
            else:
                print("取消刪除。")
        else:
            print(f"族群 '{group_name}' 不存在。")

    elif choice == "2":
        group_name = input("輸入要刪除個股的族群名稱: ").strip()
        if not group_name:
            print("族群名稱不能為空。")
            delete_group_or_stock(groups)
        if group_name in groups:
            current_stocks = groups[group_name]
            print(f"\n==============================")
            print(f"族群 '{group_name}' 中目前的個股:")
            if current_stocks:
                for idx, stock in enumerate(current_stocks, start=1):
                    print(f"  {str(idx).rjust(2)}. {stock}")
            else:
                print("  無")
            print("==============================\n")

            if not current_stocks:
                print(f"族群 '{group_name}' 中目前沒有任何個股。")
                delete_group_or_stock(groups)

            print(f"開始刪除個股從族群 '{group_name}'。")
            print("請輸入要刪除的個股代號，輸入 'Q' 以退出刪除模式。")

            while True:
                stock_name = input("輸入個股代號 (或 'Q' 退出): ").strip()
                if stock_name.upper() == "Q":
                    print("退出刪除個股模式。")
                    break
                elif not stock_name:
                    print("輸入無效，請重新輸入。")
                    continue
                elif stock_name not in groups[group_name]:
                    print(f"個股 '{stock_name}' 不存在於族群 '{group_name}' 中。")
                else:
                    confirm = input(f"確定要刪除個股 '{stock_name}' 嗎？ (Y/N): ").strip().upper()
                    if confirm == "Y":
                        groups[group_name].remove(stock_name)
                        print(f"個股 '{stock_name}' 已從族群 '{group_name}' 中刪除。")
                        if not groups[group_name]:
                            print(f"族群 '{group_name}' 現在已經沒有任何個股。")
                    else:
                        print("取消刪除。")
        else:
            print(f"族群 '{group_name}' 不存在。")

    elif choice == "0":
        print("返回主選單。")
        manage_groups()

    else:
        print("無效的選項，請重新選擇。")

def main():
    load_settings()
    config = load_config("config.yaml")
    client = RestClient(api_key=config['api_key'])
    matrix_dict_analysis = load_matrix_dict_analysis()
    main_menu()

if __name__ == "__main__":
    check_and_install_packages(required_packages)
    print("所有必要套件已安裝，開始執行程式...")
    main()